"""
pipe_report_writer.py — 管网概算 Excel 报告生成 (extracted from report_writer.py)
"""

"""

report_writer.py — 工程概算完整 Excel 报告生成



输出多 sheet 报告,按工艺流程排序:

  1. 编制说明 — 工程概况、编制依据、造价指标

  2. 总概算汇总 — 各项费用 + 按构筑物分项汇总

  3~11. 各构筑物独立明细 sheet(管网→调节池→...→紫外消毒)

  12. 施工措施费明细

  13. 设备购置概算表

  14. 其他费用明细表

  15. 主要材料汇总表(按构筑物分列)

  16. 造价指标分析

"""


import os

from _paths import setup_import_paths
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

from .pipe_network_cost import PipeNetworkEstimate
from .unit_prices import get_pipe_price

# 维度标签解析


setup_import_paths()

from _logging import get_logger

_log = get_logger(__name__)


def write_pipe_network_report(
    est: PipeNetworkEstimate,
    output_dir: str = "output",
    filename: str = "guanwang_gaisuan.xlsx",
    source_name: str = "",
) -> str:
    """生成管网工程独立概算报告(4 sheet Excel)



    Sheet 1: 编制说明

    Sheet 2: 管网概算汇总

    Sheet 3: 逐段管道造价明细

    Sheet 4: 按管径汇总



    Args:

        source_name: 管网Excel文件名(不含扩展名),用于生成输出文件名



    Returns:

        输出文件完整路径,失败返回空字符串

    """

    try:

        import openpyxl  # noqa: F811

    except ImportError:

        return ""

    os.makedirs(output_dir, exist_ok=True)

    if filename == "guanwang_gaisuan.xlsx":

        from datetime import datetime

        ts = datetime.now().strftime("%Y%m%d_%H%M%S")

        prefix = f"{source_name}_" if source_name else ""

        filename = f"{prefix}guanwang_gaisuan_{ts}.xlsx"

    fp = os.path.join(output_dir, filename)

    wb = openpyxl.Workbook()

    # ── 样式定义 ──

    TITLE_FONT = Font(name="宋体", size=16, bold=True)

    H2_FONT = Font(name="宋体", size=12, bold=True)

    HDR_FONT = Font(name="宋体", size=10, bold=True)

    NML_FONT = Font(name="宋体", size=10)

    SMALL_FONT = Font(name="宋体", size=9)

    THIN = Border(
        left=Side("thin"), right=Side("thin"), top=Side("thin"), bottom=Side("thin")
    )

    HDR_FILL = PatternFill("solid", fgColor="D9E1F2")

    SUM_FILL = PatternFill("solid", fgColor="FFF2CC")

    CENTER = Alignment(horizontal="center", vertical="center", wrap_text=True)

    LEFT = Alignment(horizontal="left", vertical="center", wrap_text=True)

    def style_header(ws, row, cols):

        for c in range(1, cols + 1):

            cell = ws.cell(row=row, column=c)

            cell.font = HDR_FONT
            cell.fill = HDR_FILL
            cell.border = THIN

            cell.alignment = CENTER

    def style_row(ws, row, cols, is_sum=False):

        for c in range(1, cols + 1):

            cell = ws.cell(row=row, column=c)

            cell.font = NML_FONT
            cell.border = THIN

            if is_sum:
                cell.fill = SUM_FILL

    # ═══════════════════════════════════════

    # Sheet 1: 编制说明

    # ═══════════════════════════════════════

    ws0 = wb.active

    ws0.title = "编制说明"

    ws0.merge_cells("A1:D1")

    c = ws0["A1"]
    c.value = "管网工程概算报告"
    c.font = TITLE_FONT

    c.alignment = CENTER

    ws0.row_dimensions[1].height = 36

    unit_cost = est.grand_total / est.total_length if est.total_length > 0 else 0

    notes = [
        ("工程名称", "污水处理厂 — 排水管网"),
        ("编制依据", "2019地方定额(2024调整) + GB50500-2013"),
        ("价格水平", "2024年"),
        ("管网总造价", f"{est.grand_total_wan:,.2f} 万元"),
        ("单位造价", f"{unit_cost:,.0f} 元/m"),
        ("", ""),
        ("分项造价", ""),
        ("  管道材料+铺设", f"{est.total_pipe_cost/10000:,.2f} 万元"),
        ("  土方开挖回填", f"{est.total_earthwork/10000:,.2f} 万元"),
        ("  施工机械台班", f"{est.total_machine/10000:,.2f} 万元"),
        ("  人工工时", f"{est.total_labor/10000:,.2f} 万元"),
        ("  检查井", f"{est.total_manhole/10000:,.2f} 万元"),
    ]

    if est.pump_station_cost > 0:

        notes.append(("  提升泵站", f"{est.pump_station_cost/10000:,.2f} 万元"))

    notes.append(("  管道段数", f"{len(est.segments)} 段"))

    for i, (label, value) in enumerate(notes, 3):

        ws0.cell(row=i, column=1, value=label).font = Font(
            name="宋体", size=10, bold=True
        )

        ws0.cell(row=i, column=2, value=str(value)).font = NML_FONT

    ws0.column_dimensions["A"].width = 22
    ws0.column_dimensions["B"].width = 50

    # ═══════════════════════════════════════

    # Sheet 2: 管网概算汇总

    # ═══════════════════════════════════════

    ws1 = wb.create_sheet("管网概算汇总")

    ws1.merge_cells("A1:E1")

    ws1["A1"].value = "管网工程概算汇总表"
    ws1["A1"].font = H2_FONT

    ws1["A1"].alignment = CENTER

    ws1.row_dimensions[1].height = 28

    headers = ["序号", "费用名称", "金额(万元)", "占比(%)", "备注"]

    for ci, h in enumerate(headers, 1):

        ws1.cell(row=2, column=ci, value=h)

    style_header(ws1, 2, 5)

    gt_wan = est.grand_total_wan if est.grand_total_wan > 0 else 1

    sum_items = [
        ("一", "管道材料+铺设", est.total_pipe_cost / 10000, "含材料及铺设人工机械"),
        ("二", "土方开挖回填", est.total_earthwork / 10000, "含放坡、工作面、余土外运"),
        ("三", "施工机械台班", est.total_machine / 10000, "起重机+载重车+挖掘机"),
        ("四", "人工工时", est.total_labor / 10000, "技工+普工,综合128元/工日"),
        ("五", "检查井摊销", est.total_manhole / 10000, "按40m间距布置"),
    ]

    if est.pump_station_cost > 0:

        sum_items.append(
            ("六", "污水提升泵站", est.pump_station_cost / 10000, "含基础+扬程附加")
        )

    sum_items.append(("", "合  计", est.grand_total_wan, ""))

    for i, (seq, name, val, note) in enumerate(sum_items, 3):

        ws1.cell(row=i, column=1, value=seq)

        ws1.cell(row=i, column=2, value=name)

        c3 = ws1.cell(row=i, column=3, value=val)

        c3.number_format = "#,##0.00"

        pct = val / gt_wan * 100 if gt_wan > 0 else 0

        ws1.cell(row=i, column=4, value=f"{pct:.1f}%" if name != "合  计" else "100%")

        ws1.cell(row=i, column=5, value=note)

        is_sum = name == "合  计"

        style_row(ws1, i, 5, is_sum)

    for i, w in enumerate([6, 22, 16, 12, 35], 1):

        ws1.column_dimensions[get_column_letter(i)].width = w

    # ═══════════════════════════════════════

    # Sheet 3: 逐段管道造价明细

    # ═══════════════════════════════════════

    ws2 = wb.create_sheet("逐段管道造价明细")

    ws2.merge_cells("A1:L1")

    ws2["A1"].value = "逐段管道造价明细表"
    ws2["A1"].font = H2_FONT

    ws2["A1"].alignment = CENTER

    ws2.row_dimensions[1].height = 28

    seg_headers = [
        "序号",
        "起点编号",
        "终点编号",
        "管径(mm)",
        "长度(m)",
        "平均埋深(m)",
        "管道费(元)",
        "土方费(元)",
        "机械费(元)",
        "人工费(元)",
        "检查井费(元)",
        "合计(元)",
    ]

    for ci, h in enumerate(seg_headers, 1):

        ws2.cell(row=2, column=ci, value=h)

    style_header(ws2, 2, 12)

    total_pipe = 0
    total_earth = 0
    total_mach = 0

    total_labor = 0
    total_mh = 0
    total_all = 0

    row = 3

    for seg in est.segments:

        avg_d = (seg.depth_start + seg.depth_end) / 2.0

        vals = [
            seg.seq,
            seg.start_node,
            seg.end_node,
            seg.diameter,
            round(seg.length, 1),
            round(avg_d, 2),
            round(seg.pipe_cost, 0),
            round(seg.earthwork_cost, 0),
            round(seg.machine_cost, 0),
            round(seg.labor_cost, 0),
            round(seg.manhole_cost, 0),
            round(seg.total, 0),
        ]

        for ci, v in enumerate(vals, 1):

            cell = ws2.cell(row=row, column=ci, value=v)

            if ci in (7, 8, 9, 10, 11, 12):

                cell.number_format = "#,##0"

            elif ci in (5, 6):

                cell.number_format = "#,##0.0" if ci == 5 else "#,##0.00"

        style_row(ws2, row, 12)

        total_pipe += seg.pipe_cost
        total_earth += seg.earthwork_cost

        total_mach += seg.machine_cost
        total_labor += seg.labor_cost

        total_mh += seg.manhole_cost
        total_all += seg.total

        row += 1

    # 合计行

    sum_vals = [
        "",
        "",
        "",
        "",
        "",
        "",
        round(total_pipe, 0),
        round(total_earth, 0),
        round(total_mach, 0),
        round(total_labor, 0),
        round(total_mh, 0),
        round(total_all, 0),
    ]

    ws2.cell(row=row, column=1, value="合计").font = HDR_FONT

    for ci in range(2, 7):

        ws2.cell(row=row, column=ci, value="")

    for ci, v in enumerate(sum_vals[6:], 7):

        cell = ws2.cell(row=row, column=ci, value=v)

        cell.number_format = "#,##0"

    style_row(ws2, row, 12, is_sum=True)

    seg_widths = [6, 10, 10, 10, 10, 12, 14, 14, 14, 14, 14, 14]

    for i, w in enumerate(seg_widths, 1):

        ws2.column_dimensions[get_column_letter(i)].width = w

    # ═══════════════════════════════════════

    # Sheet 4: 按管径汇总

    # ═══════════════════════════════════════

    ws3 = wb.create_sheet("按管径汇总")

    ws3.merge_cells("A1:F1")

    ws3["A1"].value = "按管径分类汇总表"
    ws3["A1"].font = H2_FONT

    ws3["A1"].alignment = CENTER

    ws3.row_dimensions[1].height = 28

    dia_headers = [
        "序号",
        "管径(mm)",
        "总长度(m)",
        "占比(%)",
        "综合单价(元/m)",
        "合价(万元)",
    ]

    for ci, h in enumerate(dia_headers, 1):

        ws3.cell(row=2, column=ci, value=h)

    style_header(ws3, 2, 6)

    total_len = est.total_length if est.total_length > 0 else 1

    row = 3
    seq = 0

    for d in sorted(est.by_diameter.keys()):

        seq += 1

        length = est.by_diameter[d]

        price = get_pipe_price(d)

        subtotal = length * price / 10000  # 元→万元

        ws3.cell(row=row, column=1, value=seq)

        ws3.cell(row=row, column=2, value=f"DN{d}")

        c3 = ws3.cell(row=row, column=3, value=round(length, 1))

        c3.number_format = "#,##0.0"

        ws3.cell(row=row, column=4, value=f"{length/total_len*100:.1f}%")

        c5 = ws3.cell(row=row, column=5, value=price)

        c5.number_format = "#,##0"

        c6 = ws3.cell(row=row, column=6, value=round(subtotal, 2))

        c6.number_format = "#,##0.00"

        style_row(ws3, row, 6)

        row += 1

    # 合计行

    total_wan_from_dia = sum(
        est.by_diameter[d] * get_pipe_price(d) / 10000 for d in est.by_diameter
    )

    ws3.cell(row=row, column=1, value="").font = NML_FONT

    ws3.cell(row=row, column=2, value="合计").font = HDR_FONT

    c3 = ws3.cell(row=row, column=3, value=round(est.total_length, 1))

    c3.number_format = "#,##0.0"

    ws3.cell(row=row, column=4, value="100%")

    ws3.cell(row=row, column=5, value="")

    c6 = ws3.cell(row=row, column=6, value=round(total_wan_from_dia, 2))

    c6.number_format = "#,##0.00"

    style_row(ws3, row, 6, is_sum=True)

    dia_widths = [6, 12, 14, 10, 16, 14]

    for i, w in enumerate(dia_widths, 1):

        ws3.column_dimensions[get_column_letter(i)].width = w

    wb.save(fp)

    return fp

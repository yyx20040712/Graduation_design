"""
report_writer.py — 工程概算完整 Excel 报告生成

输出多 sheet 报告,按工艺流程排序:
  1. 编制说明 — 工程概况、编制依据、造价指标
  2. 总概算汇总 — 各项费用 + 按构筑物分项汇总
  3~11. 各构筑物独立明细 sheet(管网→调节池→...→紫外消毒)
  12. 施工措施费明细
  13. 设备购置概算表
  14. 其他费用明细表
  15. 主要材料汇总表(按构筑物分列)
  16. 造价指标分析
"""

import os
from typing import Optional

# 维度标签解析
from _paths import setup_import_paths
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

from .cost_estimator import EstimateResult
from .unit_prices import INDIRECT_RATES

setup_import_paths()
from _logging import get_logger
from ui.dimension_labels import (
    format_dimension_row,
    is_internal_debug_dim,
    is_water_quality_dim,
)

_log = get_logger(__name__)


# ── 工艺流程顺序 (node_type → 显示名称) ──
def _build_flow_order():
    """Auto-derive flow order from ModManager's process_stage grouping."""
    try:
        from mods.mod_manager import get_mod_manager

        mgr = get_mod_manager()
        mgr.discover_all()
        order = mgr.get_flow_order()
        if order:
            return order
    except Exception as e:
        _log.warning("operation failed: %s", e, exc_info=True)
    # Legacy fallback
    return [
        ("pipe_network", "管网工程"),
        ("tiaojiechi", "调节池"),
        ("cugeshan", "粗格栅"),
        ("xigeshan", "细格栅"),
        ("chenshachi", "旋流沉砂池"),
        ("chuchenchi", "辐流初沉池"),
        ("cass", "CASS反应器"),
        ("aao", "AAO反应池"),
        ("gaomidu", "高密度沉淀池"),
        ("vxinglvchi", "V型滤池"),
        ("ziwai", "紫外消毒池"),
        ("kw_tiaojiechi", "矿井水调节池"),
        ("kw_chenshachi", "平流沉砂池"),
        ("kw_ningjiao", "混凝反应池"),
        ("kw_cifenli", "磁分离"),
    ]


FLOW_ORDER = _build_flow_order()


def _safe_name(name: str) -> str:
    """截断至 31 字符"""
    return name[:31]


def write_cost_report(
    est: EstimateResult,
    executor=None,
    output_dir: str = "output",
    filename: str = "gaisuan_baogao.xlsx",
    results: Optional[dict] = None,
) -> str:
    """生成详细工程概算报告(按工艺流程排序,每构筑物独立 sheet)

    Args:
        est: EstimateResult 对象(含 node_type 信息的 BOQItem 列表)
        executor: GraphExecutor 实例(可选,用于获取节点尺寸参数)
        output_dir: 输出目录
        filename: 文件名
        results: executor.execute() 返回的 {node_id: NodeResult} 字典.
                 提供时优先使用, 避免读过期 node._result 缓存.

    Returns:
        输出文件完整路径,失败返回空字符串
    """
    try:
        import openpyxl  # noqa: F811
    except ImportError:
        return ""

    os.makedirs(output_dir, exist_ok=True)
    if filename == "gaisuan_baogao.xlsx":
        from datetime import datetime

        ts = datetime.now().strftime("%Y%m%d_%H%M%S")
        filename = f"gaisuan_baogao_{ts}.xlsx"
    fp = os.path.join(output_dir, filename)
    wb = openpyxl.Workbook()

    # ── 样式 ──
    TITLE_FONT = Font(name="宋体", size=16, bold=True)
    H2_FONT = Font(name="宋体", size=12, bold=True)
    HDR_FONT = Font(name="宋体", size=10, bold=True)
    NML_FONT = Font(name="宋体", size=10)
    SMALL_FONT = Font(name="宋体", size=9)
    THIN = Border(
        left=Side("thin"), right=Side("thin"), top=Side("thin"), bottom=Side("thin")
    )
    HDR_FILL = PatternFill("solid", fgColor="D9E1F2")
    SUM_FILL = PatternFill("solid", fgColor="FFF2CC")
    CENTER = Alignment(horizontal="center", vertical="center", wrap_text=True)
    LEFT = Alignment(horizontal="left", vertical="center", wrap_text=True)

    def style_header(ws, row, cols):
        for c in range(1, cols + 1):
            cell = ws.cell(row=row, column=c)
            cell.font = HDR_FONT
            cell.fill = HDR_FILL
            cell.border = THIN
            cell.alignment = CENTER

    def style_row(ws, row, cols, is_sum=False):
        for c in range(1, cols + 1):
            cell = ws.cell(row=row, column=c)
            cell.font = NML_FONT
            cell.border = THIN
            if is_sum:
                cell.fill = SUM_FILL

    # ═══════════════════════════════════════
    # Sheet 1: 编制说明
    # ═══════════════════════════════════════
    ws0 = wb.active
    ws0.title = "编制说明"
    ws0.merge_cells("A1:D1")
    c = ws0["A1"]
    c.value = f"{est.project_name} 工程概算报告"
    c.font = TITLE_FONT
    c.alignment = CENTER
    ws0.row_dimensions[1].height = 36

    # 按 node_type 汇总各构筑物造价
    struct_costs = {}
    for item in est.items:
        if item.node_type and item.category == "建筑工程":
            struct_costs[item.node_type] = (
                struct_costs.get(item.node_type, 0) + item.total
            )

    notes = [
        ("工程名称", est.project_name),
        ("处理规模", f"{est.q_daily:.0f} m³/d" if est.q_daily > 0 else "—"),
        ("出水标准", "GB18918-2002 一级A"),
        ("编制日期", datetime.now().strftime("%Y年%m月%d日")),
        ("编制依据", "2019地方定额(2024调整) + GB50500-2013 + T/BCEBCA1-2023"),
        ("价格水平", "2024年"),
        ("", ""),
        ("═══ 造价总览 ═══", ""),
        ("概算总造价", f"{est.total_cost:,.2f} 万元"),
        ("单位造价", f"{est.unit_cost:,.0f} 元/(m³·d)" if est.unit_cost else "—"),
        ("造价校验", est.check_msg if est.check_msg else "—"),
        ("", ""),
        ("═══ 分项费用 ═══", ""),
        ("建筑工程费", f"{est.civil_cost:,.2f} 万元"),
        ("设备购置费", f"{est.equip_cost:,.2f} 万元"),
        ("安装工程费", f"{est.install_cost:,.2f} 万元"),
        ("其他费用", f"{est.other_cost:,.2f} 万元"),
        ("预备费", f"{est.contingency:,.2f} 万元"),
        ("", ""),
        ("═══ 各构筑物土建造价 ═══", ""),
    ]
    for nt, dname in FLOW_ORDER:
        if nt in struct_costs:
            notes.append((f"  {dname}", f"{struct_costs[nt]/10000:,.2f} 万元"))

    for i, (label, value) in enumerate(notes, 3):
        ws0.cell(row=i, column=1, value=label).font = Font(
            name="宋体", size=10, bold=True
        )
        ws0.cell(row=i, column=2, value=str(value)).font = NML_FONT
    ws0.column_dimensions["A"].width = 22
    ws0.column_dimensions["B"].width = 55

    # ═══════════════════════════════════════
    # Sheet 2: 总概算汇总
    # ═══════════════════════════════════════
    ws1 = wb.create_sheet("总概算汇总")
    ws1.merge_cells("A1:F1")
    ws1["A1"].value = "总概算汇总表"
    ws1["A1"].font = H2_FONT
    ws1["A1"].alignment = CENTER
    ws1.row_dimensions[1].height = 28

    headers = ["序号", "费用名称", "金额(万元)", "占比(%)", "计算基数", "备注"]
    for ci, h in enumerate(headers, 1):
        ws1.cell(row=2, column=ci, value=h)
    style_header(ws1, 2, 6)

    sums = [
        ("一", "建筑工程费", est.civil_cost, "—", "含土建+管网+措施"),
        ("二", "设备购置费", est.equip_cost, "—", "含工艺设备+通用设备"),
        ("三", "安装工程费", est.install_cost, f"设备费×{15}%", ""),
        ("四", "其他费用", est.other_cost, f"建安费×{12.5}%", "管理+设计+监理+前期"),
        ("五", "预备费", est.contingency, "前四项×10%", "基本预备费"),
        (
            "六",
            "增值税",
            round(
                est.total_cost
                - sum(
                    [
                        est.civil_cost,
                        est.equip_cost,
                        est.install_cost,
                        est.other_cost,
                        est.contingency,
                    ]
                ),
                2,
            ),
            "前五项×9%",
            "建筑业增值税",
        ),
        ("", "工程总造价(含税)", est.total_cost, "100%", ""),
    ]
    for i, (seq, name, val, base, note) in enumerate(sums, 3):
        ws1.cell(row=i, column=1, value=seq)
        ws1.cell(row=i, column=2, value=name)
        c3 = ws1.cell(row=i, column=3, value=val)
        c3.number_format = "#,##0.00"
        pct = val / est.total_cost * 100 if est.total_cost else 0
        ws1.cell(
            row=i,
            column=4,
            value=f"{pct:.1f}%" if name != "工程总造价(含税)" else "100%",
        )
        ws1.cell(row=i, column=5, value=base)
        ws1.cell(row=i, column=6, value=note)
        is_sum = name == "工程总造价(含税)"
        style_row(ws1, i, 6, is_sum)

    # ── 建筑工程费分项 ──
    row = i + 2
    ws1.merge_cells(start_row=row, start_column=1, end_row=row, end_column=6)
    ws1.cell(row=row, column=1, value="▎建筑工程费 — 按构筑物分项").font = H2_FONT
    style_row(ws1, row, 6)
    row += 1
    for ci, h in enumerate(["", "构筑物", "土建造价(万元)", "占建筑费%", "", ""], 1):
        ws1.cell(row=row, column=ci, value=h)
    style_header(ws1, row, 6)
    row += 1
    civil_total = est.civil_cost if est.civil_cost > 0 else 1
    seq2 = 0
    for nt, dname in FLOW_ORDER:
        cost = struct_costs.get(nt, 0) / 10000
        if cost <= 0:
            continue
        seq2 += 1
        ws1.cell(row=row, column=1, value=seq2)
        ws1.cell(row=row, column=2, value=dname)
        c3 = ws1.cell(row=row, column=3, value=cost)
        c3.number_format = "#,##0.00"
        ws1.cell(row=row, column=4, value=f"{cost/est.civil_cost*100:.1f}%")
        style_row(ws1, row, 6)
        row += 1

    for i, w in enumerate([6, 14, 16, 12, 16, 26], 1):
        ws1.column_dimensions[get_column_letter(i)].width = w

    # ═══════════════════════════════════════
    # Sheets 3~11: 各构筑物独立明细
    # ═══════════════════════════════════════
    for nt, dname in FLOW_ORDER:
        civil_items = [
            it for it in est.items if it.node_type == nt and it.category == "建筑工程"
        ]
        equip_items = [
            it for it in est.items if it.node_type == nt and it.category == "设备购置"
        ]
        if not civil_items and not equip_items:
            continue

        _write_structure_sheet(
            wb,
            dname,
            nt,
            civil_items,
            equip_items,
            executor,
            H2_FONT,
            HDR_FONT,
            NML_FONT,
            SMALL_FONT,
            THIN,
            HDR_FILL,
            SUM_FILL,
            CENTER,
            LEFT,
            results=results,
        )

    # ═══════════════════════════════════════
    # Sheet: 施工措施费明细
    # ═══════════════════════════════════════
    measure_items = [
        it
        for it in est.items
        if it.code.startswith("CS-") and it.category == "建筑工程"
    ]
    if measure_items:
        ws_m = wb.create_sheet("施工措施费明细")
        ws_m.merge_cells("A1:F1")
        ws_m["A1"].value = "施工措施费明细表"
        ws_m["A1"].font = H2_FONT
        ws_m["A1"].alignment = CENTER
        ws_m.row_dimensions[1].height = 28

        mh = ["序号", "措施项目", "计算方式", "费率/单价", "金额(万元)", "备注"]
        for ci, h in enumerate(mh, 1):
            ws_m.cell(row=2, column=ci, value=h)
        style_header(ws_m, 2, 6)

        notes_map = {
            "CS-1": ("施工降水(井点降水)", "建筑工程费", "2%", "按埋深>3m基坑"),
            "CS-2": ("场地准备及临时设施", "建筑工程费", "5%", "道路/围挡/临建"),
            "CS-3": ("施工环保措施", "建筑工程费", "1.5%", "扬尘/噪声/废水"),
            "CS-4": ("调试与试运行", "设备购置费", "3%", "菌种培养/药剂/联调"),
        }
        row = 3
        seq = 0
        m_total = 0
        for it in measure_items:
            seq += 1
            note = notes_map.get(it.code, ("", "", "", ""))
            ws_m.cell(row=row, column=1, value=seq)
            ws_m.cell(row=row, column=2, value=note[0])
            ws_m.cell(row=row, column=3, value=note[1])
            ws_m.cell(row=row, column=4, value=note[2])
            c5 = ws_m.cell(row=row, column=5, value=it.total / 10000)
            c5.number_format = "#,##0.00"
            ws_m.cell(row=row, column=6, value=note[3])
            style_row(ws_m, row, 6)
            m_total += it.total
            row += 1
        # 小计
        ws_m.cell(row=row, column=2, value="措施费小计").font = HDR_FONT
        c5 = ws_m.cell(row=row, column=5, value=m_total / 10000)
        c5.number_format = "#,##0.00"
        style_row(ws_m, row, 6, is_sum=True)
        for i, w in enumerate([6, 24, 16, 12, 16, 24], 1):
            ws_m.column_dimensions[get_column_letter(i)].width = w

    # ═══════════════════════════════════════
    # Sheet: 设备购置明细
    # ═══════════════════════════════════════
    equip_items_all = [it for it in est.items if it.category == "设备购置"]
    if equip_items_all:
        ws_e = wb.create_sheet("设备购置明细")
        ws_e.merge_cells("A1:F1")
        ws_e["A1"].value = "设备购置明细表"
        ws_e["A1"].font = H2_FONT
        ws_e["A1"].alignment = CENTER
        ws_e.row_dimensions[1].height = 28

        eh = ["序号", "设备名称", "所属构筑物", "数量", "单价(万元)", "合价(万元)"]
        for ci, h in enumerate(eh, 1):
            ws_e.cell(row=2, column=ci, value=h)
        style_header(ws_e, 2, 6)

        row = 3
        seq = 0
        e_total = 0
        for it in equip_items_all:
            seq += 1
            # 查找 node_type 对应的中文名
            struct_name = ""
            for nt, dn in FLOW_ORDER:
                if it.node_type == nt:
                    struct_name = dn
                    break
            if not struct_name and it.code == "SB-COM":
                struct_name = "全厂通用"

            ws_e.cell(row=row, column=1, value=seq)
            # 设备名称: 去掉 "构筑物名—" 前缀
            ename = it.name
            if "—" in ename:
                ename = ename.split("—", 1)[1]
            ws_e.cell(row=row, column=2, value=ename)
            ws_e.cell(row=row, column=3, value=struct_name)
            ws_e.cell(row=row, column=4, value=it.quantity)
            c5 = ws_e.cell(row=row, column=5, value=it.unit_price / 10000)
            c5.number_format = "#,##0.00"
            c6 = ws_e.cell(row=row, column=6, value=it.total / 10000)
            c6.number_format = "#,##0.00"
            style_row(ws_e, row, 6)
            e_total += it.total
            row += 1
        # 小计
        ws_e.cell(row=row, column=2, value="设备费小计").font = HDR_FONT
        c6 = ws_e.cell(row=row, column=6, value=e_total / 10000)
        c6.number_format = "#,##0.00"
        style_row(ws_e, row, 6, is_sum=True)
        for i, w in enumerate([6, 28, 14, 8, 14, 14], 1):
            ws_e.column_dimensions[get_column_letter(i)].width = w

    # ═══════════════════════════════════════
    # Sheet: 其他费用明细
    # ═══════════════════════════════════════
    ws4 = wb.create_sheet("其他费用明细")
    ws4.merge_cells("A1:D1")
    ws4["A1"].value = "其他费用明细表"
    ws4["A1"].font = H2_FONT
    ws4["A1"].alignment = CENTER

    base_ca = est.civil_cost + est.install_cost
    other_items = [
        (
            "建设单位管理费",
            base_ca * INDIRECT_RATES["management"],
            f"{INDIRECT_RATES['management']*100:.0f}%",
            "建安费",
        ),
        (
            "勘察设计费",
            base_ca * INDIRECT_RATES["design"],
            f"{INDIRECT_RATES['design']*100:.0f}%",
            "建安费",
        ),
        (
            "工程监理费",
            base_ca * INDIRECT_RATES["supervision"],
            f"{INDIRECT_RATES['supervision']*100:.1f}%",
            "建安费",
        ),
        (
            "前期工作费",
            base_ca * INDIRECT_RATES["preparation"],
            f"{INDIRECT_RATES['preparation']*100:.0f}%",
            "建安费",
        ),
    ]
    for ci, h in enumerate(["序号", "费用名称", "金额(万元)", "费率", "计算基数"], 1):
        ws4.cell(row=2, column=ci, value=h)
    style_header(ws4, 2, 5)
    for i, (name, val, rate, base) in enumerate(other_items, 3):
        ws4.cell(row=i, column=1, value=i - 2)
        ws4.cell(row=i, column=2, value=name)
        c3 = ws4.cell(row=i, column=3, value=val)
        c3.number_format = "#,##0.00"
        ws4.cell(row=i, column=4, value=rate)
        ws4.cell(row=i, column=5, value=base)
        style_row(ws4, i, 5)
    for i, w in enumerate([6, 22, 16, 10, 10], 1):
        ws4.column_dimensions[get_column_letter(i)].width = w

    # ═══════════════════════════════════════
    # Sheet: 主要材料汇总(按构筑物分列)
    # ═══════════════════════════════════════
    ws5 = wb.create_sheet("主要材料汇总")
    ws5.merge_cells("A1:H1")
    ws5["A1"].value = "主要材料估算量 — 按构筑物分列"
    ws5["A1"].font = H2_FONT
    ws5["A1"].alignment = CENTER
    ws5.row_dimensions[1].height = 28

    # 表头: 构筑物 | 混凝土(m³) | 钢筋(t) | 土方(m³) | 防水(m²) | 模板(m²) | 垫层(m³) | 土建小计(万元)
    mh2 = [
        "构筑物",
        "混凝土(m³)",
        "钢筋(t)",
        "土方(m³)",
        "防水(m²)",
        "模板(m²)",
        "垫层(m³)",
        "土建(万元)",
    ]
    for ci, h in enumerate(mh2, 1):
        ws5.cell(row=2, column=ci, value=h)
    style_header(ws5, 2, 8)

    row = 3
    totals = [0] * 8  # 累计
    for nt, dname in FLOW_ORDER:
        items_nt = [
            it for it in est.items if it.node_type == nt and it.category == "建筑工程"
        ]
        if not items_nt:
            continue
        conc = sum(
            it.quantity
            for it in items_nt
            if it.unit == "m³" and ("C30" in it.name or "C15" in it.name)
        )
        rebar = sum(
            it.quantity for it in items_nt if it.unit == "t" and "钢筋" in it.name
        )
        excav = sum(
            it.quantity
            for it in items_nt
            if it.unit == "m³" and ("土方" in it.name or "挖" in it.name)
        )
        wp = sum(
            it.quantity for it in items_nt if it.unit == "m²" and "防水" in it.name
        )
        fm = sum(
            it.quantity for it in items_nt if it.unit == "m²" and "模板" in it.name
        )
        pad = sum(
            it.quantity for it in items_nt if it.unit == "m³" and "垫层" in it.name
        )
        cost_wan = sum(it.total for it in items_nt) / 10000

        vals = [
            dname,
            round(conc, 0),
            round(rebar, 1),
            round(excav, 0),
            round(wp, 0),
            round(fm, 0),
            round(pad, 0),
            round(cost_wan, 2),
        ]
        for ci, v in enumerate(vals, 1):
            cell = ws5.cell(row=row, column=ci, value=v)
            if ci >= 2 and ci <= 7:
                cell.number_format = "#,##0" if ci != 3 else "#,##0.0"
            elif ci == 8:
                cell.number_format = "#,##0.00"
        style_row(ws5, row, 8)
        for j in range(8):
            if isinstance(vals[j], (int, float)):
                totals[j] += vals[j]
        row += 1

    # 合计行
    totals[0] = "合计"
    for ci, v in enumerate(totals, 1):
        cell = ws5.cell(row=row, column=ci, value=v if ci > 1 else "合计")
        if ci == 3:
            cell.number_format = "#,##0.0"
        elif ci >= 2:
            cell.number_format = "#,##0" if ci <= 7 else "#,##0.00"
    style_row(ws5, row, 8, is_sum=True)

    for i, w in enumerate([14, 14, 10, 12, 12, 12, 10, 14], 1):
        ws5.column_dimensions[get_column_letter(i)].width = w

    # ═══════════════════════════════════════
    # Sheet: 造价指标分析
    # ═══════════════════════════════════════
    ws6 = wb.create_sheet("造价指标分析")
    ws6.merge_cells("A1:D1")
    ws6["A1"].value = "造价指标分析"
    ws6["A1"].font = H2_FONT
    ws6["A1"].alignment = CENTER
    ws6.row_dimensions[1].height = 28

    indicators = [
        ("处理规模", f"{est.q_daily:.0f} m³/d" if est.q_daily > 0 else "—"),
        ("工程总造价", f"{est.total_cost:,.2f} 万元"),
        (
            "单位处理规模造价",
            f"{est.unit_cost:,.0f} 元/(m³·d)" if est.unit_cost else "—",
        ),
        ("", ""),
        ("造价构成分析", ""),
        (
            "  建筑工程费占比",
            f"{est.civil_cost/est.total_cost*100:.1f}%" if est.total_cost else "—",
        ),
        (
            "  设备购置费占比",
            f"{est.equip_cost/est.total_cost*100:.1f}%" if est.total_cost else "—",
        ),
        (
            "  安装工程费占比",
            f"{est.install_cost/est.total_cost*100:.1f}%" if est.total_cost else "—",
        ),
        (
            "  其他费用占比",
            f"{est.other_cost/est.total_cost*100:.1f}%" if est.total_cost else "—",
        ),
        (
            "  预备费占比",
            f"{est.contingency/est.total_cost*100:.1f}%" if est.total_cost else "—",
        ),
        ("", ""),
        ("参考标准", "T/BCEBCA 1-2023 一级A: 3000~5000 元/(m³·d)"),
        ("合理性校验", est.check_msg if est.check_msg else "—"),
    ]
    for i, (label, value) in enumerate(indicators, 3):
        ws6.cell(row=i, column=1, value=label).font = Font(
            name="宋体", size=10, bold=True
        )
        ws6.cell(row=i, column=2, value=str(value)).font = NML_FONT
    ws6.column_dimensions["A"].width = 22
    ws6.column_dimensions["B"].width = 55

    wb.save(fp)
    return fp


def _write_structure_sheet(
    wb,
    dname,
    nt,
    civil_items,
    equip_items,
    executor,
    H2_FONT,
    HDR_FONT,
    NML_FONT,
    SMALL_FONT,
    THIN,
    HDR_FILL,
    SUM_FILL,
    CENTER,
    LEFT,
    results: Optional[dict] = None,
):
    """写入单个构筑物的独立明细 sheet"""
    ws = wb.create_sheet(_safe_name(dname))

    # ── 标题 ──
    ws.merge_cells("A1:G1")
    ws["A1"].value = f"{dname} — 工程概算明细"
    ws["A1"].font = H2_FONT
    ws["A1"].alignment = CENTER
    ws.row_dimensions[1].height = 28

    row = 3

    # ── 设计参数与尺寸(如果有 executor)──
    if executor and nt != "pipe_network":
        node = None
        for nid, n in executor._nodes.items():
            if n.NODE_TYPE == nt:
                node = n
                break
        if node:
            # 优先使用 results 字典中的结果, 回退到 node.result
            res = results.get(node.node_id) if results else None
            if res is None:
                res = node.result
        else:
            res = None
        if res and res.success:
            dims = dict(res.dimensions)
            params = res.params or {}
            if dims or params:
                ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=7)
                ws.cell(row=row, column=1, value="▎设计参数与构筑物尺寸").font = Font(
                    name="宋体", size=10, bold=True
                )
                ws.cell(row=row, column=1).fill = HDR_FILL
                for c in range(1, 8):
                    ws.cell(row=row, column=c).border = THIN
                row += 1

                # 4列表头
                for ci, h in enumerate(["符号", "物理意义", "单位", "取值"], 1):
                    c = ws.cell(row=row, column=ci, value=h)
                    c.font = HDR_FONT
                    c.fill = HDR_FILL
                    c.border = THIN
                    c.alignment = CENTER
                row += 1

                # 设计参数
                for k, v in params.items():
                    sym, meaning, default_u = format_dimension_row(k, v, "")
                    val_str = f"{v:.2f}" if isinstance(v, float) else str(v)
                    ws.cell(row=row, column=1, value=sym).font = NML_FONT
                    ws.cell(row=row, column=2, value=meaning).font = NML_FONT
                    ws.cell(row=row, column=3, value=default_u).font = NML_FONT
                    ws.cell(row=row, column=4, value=val_str).font = NML_FONT
                    for c in range(1, 5):
                        ws.cell(row=row, column=c).border = THIN
                    row += 1

                # 构筑物尺寸 (排除水质相关)
                for dk, (dv, du) in dims.items():
                    if is_water_quality_dim(dk, du) or is_internal_debug_dim(dk):
                        continue
                    sym, meaning, default_u = format_dimension_row(dk, dv, du)
                    unit = du or default_u
                    val_str = f"{dv:.2f}" if isinstance(dv, float) else str(dv)
                    ws.cell(row=row, column=1, value=sym).font = NML_FONT
                    ws.cell(row=row, column=2, value=meaning).font = NML_FONT
                    ws.cell(row=row, column=3, value=unit).font = NML_FONT
                    ws.cell(row=row, column=4, value=val_str).font = NML_FONT
                    for c in range(1, 5):
                        ws.cell(row=row, column=c).border = THIN
                    row += 1
                row += 1

    # ── 土建BOQ ──
    if civil_items:
        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=7)
        ws.cell(row=row, column=1, value="▎土建分部分项工程量清单").font = Font(
            name="宋体", size=10, bold=True
        )
        ws.cell(row=row, column=1).fill = HDR_FILL
        for c in range(1, 8):
            ws.cell(row=row, column=c).border = THIN
        row += 1

        headers = [
            "序号",
            "定额编号",
            "项目名称",
            "单位",
            "工程量",
            "综合单价(元)",
            "合价(元)",
        ]
        for ci, h in enumerate(headers, 1):
            ws.cell(row=row, column=ci, value=h)
            cc = ws.cell(row=row, column=ci)
            cc.font = HDR_FONT
            cc.fill = HDR_FILL
            cc.border = THIN
            cc.alignment = CENTER
        row += 1

        subtotal = 0
        for it in civil_items:
            ws.cell(row=row, column=1, value=it.seq)
            ws.cell(row=row, column=2, value=it.code)
            # 去掉前缀显示
            ename = it.name.split("—", 1)[1] if "—" in it.name else it.name
            ws.cell(row=row, column=3, value=ename)
            ws.cell(row=row, column=4, value=it.unit)
            ws.cell(row=row, column=5, value=it.quantity)
            c6 = ws.cell(row=row, column=6, value=it.unit_price)
            c6.number_format = "#,##0"
            c7 = ws.cell(row=row, column=7, value=it.total)
            c7.number_format = "#,##0"
            for c in range(1, 8):
                ws.cell(row=row, column=c).font = NML_FONT
                ws.cell(row=row, column=c).border = THIN
            subtotal += it.total
            row += 1

        # 小计
        ws.cell(row=row, column=3, value="土建小计").font = HDR_FONT
        c7 = ws.cell(row=row, column=7, value=subtotal)
        c7.number_format = "#,##0"
        c7.font = HDR_FONT
        for c in range(1, 8):
            ws.cell(row=row, column=c).fill = SUM_FILL
            ws.cell(row=row, column=c).border = THIN
        row += 2

    # ── 设备清单 ──
    if equip_items:
        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=7)
        ws.cell(row=row, column=1, value="▎工艺设备清单").font = Font(
            name="宋体", size=10, bold=True
        )
        ws.cell(row=row, column=1).fill = HDR_FILL
        for c in range(1, 8):
            ws.cell(row=row, column=c).border = THIN
        row += 1

        eh = [
            "序号",
            "设备名称",
            "规格/型号",
            "数量",
            "单位",
            "单价(万元)",
            "合价(万元)",
        ]
        for ci, h in enumerate(eh, 1):
            ws.cell(row=row, column=ci, value=h)
            cc = ws.cell(row=row, column=ci)
            cc.font = HDR_FONT
            cc.fill = HDR_FILL
            cc.border = THIN
            cc.alignment = CENTER
        row += 1

        esub = 0
        for it in equip_items:
            ws.cell(row=row, column=1, value=it.seq)
            ename = it.name.split("—", 1)[1] if "—" in it.name else it.name
            ws.cell(row=row, column=2, value=ename)
            ws.cell(row=row, column=3, value="详见设备表")
            ws.cell(row=row, column=4, value=it.quantity)
            ws.cell(row=row, column=5, value=it.unit)
            c6 = ws.cell(row=row, column=6, value=it.unit_price / 10000)
            c6.number_format = "#,##0.00"
            c7 = ws.cell(row=row, column=7, value=it.total / 10000)
            c7.number_format = "#,##0.00"
            for c in range(1, 8):
                ws.cell(row=row, column=c).font = NML_FONT
                ws.cell(row=row, column=c).border = THIN
            esub += it.total
            row += 1
        # 小计
        ws.cell(row=row, column=3, value="设备小计").font = HDR_FONT
        c7 = ws.cell(row=row, column=7, value=esub / 10000)
        c7.number_format = "#,##0.00"
        for c in range(1, 8):
            ws.cell(row=row, column=c).fill = SUM_FILL
            ws.cell(row=row, column=c).border = THIN
        row += 2

    # ── 本构筑物概算合计 ──
    civil_sub = sum(it.total for it in civil_items) / 10000
    equip_sub = sum(it.total for it in equip_items) / 10000
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=7)
    ws.cell(
        row=row,
        column=1,
        value=f"本构筑物合计: 土建 {civil_sub:,.2f} 万元 + 设备 {equip_sub:,.2f} 万元 = {civil_sub+equip_sub:,.2f} 万元",
    ).font = Font(name="宋体", size=10, bold=True)
    for c in range(1, 8):
        ws.cell(row=row, column=c).fill = SUM_FILL
        ws.cell(row=row, column=c).border = THIN

    widths = [6, 14, 32, 10, 14, 16, 16]
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w

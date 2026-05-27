"""
input_node.py — 输入节点

工艺流程的起点.设置初始水量、水质参数.
可手动输入,也可从 Excel 文件加载.

输入端口: 无
输出端口: 1个 MIXED 端口(水量+水质)
"""

from typing import Dict, List, Optional

from _logging import get_logger

_log = get_logger(__name__)

from .base import (
    NodeBase,
    NodeResult,
    NodeState,
    ParamDef,
    Port,
    PortType,
    WaterFlow,
    WaterQuality,
)


class InputNode(NodeBase):
    """进水输入节点 — 定义污水处理厂的进水量和进水水质"""

    NODE_TYPE = "input_node"
    NODE_NAME = "进水节点"
    NODE_CATEGORY = "输入/输出"

    @classmethod
    def _default_params(cls) -> Dict[str, float]:
        return {
            "Q_design": 0.57,  # m³/s
            "Q_avg_daily": 34760.7,  # m³/d
            "Kz": 1.4,  # 总变化系数
        }

    def _build_param_defs(self) -> List[ParamDef]:
        return [
            ParamDef(
                "设计流量 Q",
                "Q_design",
                value=0.57,
                default=0.57,
                min_val=0.01,
                max_val=10.0,
                step=0.01,
                unit="m³/s",
                description="最大设计流量",
            ),
            ParamDef(
                "平均日流量",
                "Q_avg_daily",
                value=34760.7,
                default=34760.7,
                min_val=1000,
                max_val=1000000,
                step=1,
                unit="m³/d",
                description="平均日处理水量",
            ),
            ParamDef(
                "总变化系数 Kz",
                "Kz",
                value=1.4,
                default=1.4,
                min_val=1.0,
                max_val=3.0,
                step=0.1,
                unit="",
                description="总变化系数,Kz = 总最高时流量 / 平均时流量",
            ),
        ]

    def _init_ports(self) -> None:
        """输入节点只有输出端口,没有输入端口"""
        # 不调用 super(),因为不需要输入端口
        self.input_ports = []  # 无输入
        self.output_ports = [
            Port(
                port_id=f"{self.node_id}-out",
                name="出水(水量+水质)",
                port_type=PortType.MIXED,
                direction="output",
                node_id=self.node_id,
            )
        ]

    # ── 水质参数(存储在实例变量中,不在 _params 里)──
    def __init__(self, node_id: Optional[str] = None):
        super().__init__(node_id)
        # 默认进水水质(来自中期报告表3-1:城市污水)
        self.water_quality = WaterQuality(
            BOD5=200.0,
            COD=400.0,
            SS=220.0,
            NH3N=35.0,
            TN=45.0,
            TP=5.0,
            pH=7.0,
        )

    def set_water_quality(self, **kwargs: float) -> None:
        """设置进水水质参数

        Example:
            node.set_water_quality(BOD5=180, COD=350)
        """
        for key, value in kwargs.items():
            if hasattr(self.water_quality, key):
                setattr(self.water_quality, key, value)
        self.state = NodeState.DIRTY

    def calculate(self, flow: WaterFlow, quality: WaterQuality) -> NodeResult:
        """输入节点直接返回设定的水量和水质

        flow 和 quality 参数不使用(上游无数据),
        输入节点从自身参数构建输出.
        """
        Qd = self.get_param("Q_design")
        Qad = self.get_param("Q_avg_daily")
        Kz = self.get_param("Kz")

        flow_out = WaterFlow(Q_design=Qd, Q_avg_daily=Qad, Kz=Kz)

        result = NodeResult(success=True)
        result.params = {
            "Q_design": Qd,
            "Q_avg_daily": Qad,
            "Kz": Kz,
        }
        result.add_dimension("设计流量", Qd, "m³/s")
        result.add_dimension("平均日流量", Qad, "m³/d")
        result.add_dimension("平均时流量", flow_out.Q_avg_hourly, "m³/h")
        result.add_dimension("变化系数", Kz, "")

        # 水质信息
        wq = self.water_quality
        for attr in ["BOD5", "COD", "SS", "NH3N", "TN", "TP"]:
            val = getattr(wq, attr)
            result.add_dimension(f"进水{attr}", val, "mg/L")

        return result

    def execute(self, flow: WaterFlow, quality: WaterQuality):
        """重写 execute: 输出流量和水质使用用户设定值而非上游输入

        InputNode 是工艺流程起点,其水量(Q_design/Q_avg_daily/Kz)
        和水质(water_quality)来自用户直接输入的参数,
        不应依赖上游传递的 flow/quality(上游不存在).
        """
        from .base import WaterQuality as WQ

        self.state = NodeState.COMPUTING
        try:
            result = self.calculate(flow, quality)
            self._result = result
            if result.success:
                self.state = NodeState.CLEAN
                Qd = self.get_param("Q_design")
                Qad = self.get_param("Q_avg_daily")
                Kz = self.get_param("Kz")
                out_flow = WaterFlow(Q_design=Qd, Q_avg_daily=Qad, Kz=Kz)
                downstream_quality = WQ(
                    BOD5=self.water_quality.BOD5,
                    COD=self.water_quality.COD,
                    SS=self.water_quality.SS,
                    NH3N=self.water_quality.NH3N,
                    TN=self.water_quality.TN,
                    TP=self.water_quality.TP,
                    pH=self.water_quality.pH,
                )
                result.inlet_quality = downstream_quality
                result.outlet_quality = downstream_quality
                return result, out_flow, downstream_quality
            else:
                self.state = NodeState.ERROR
                return result, flow, quality
        except Exception as e:
            self.state = NodeState.ERROR
            failed = NodeResult.failed(str(e))
            self._result = failed
            return failed, flow, quality

    # ── 序列化扩展 ──
    def to_dict(self) -> Dict:
        d = super().to_dict()
        d["water_quality"] = self.water_quality.to_dict()
        return d

    @classmethod
    def from_dict(cls, d: Dict) -> "InputNode":
        node = cls(node_id=d["id"])
        node.x = d["position"]["x"]
        node.y = d["position"]["y"]
        for key, value in d.get("params", {}).items():
            node._params[key] = value
        if "water_quality" in d:
            node.water_quality = WaterQuality.from_dict(d["water_quality"])
        return node

    # ── Excel 加载接口 ──
    def load_from_excel(self, filepath: str) -> bool:
        """从 Excel 文件加载水量和水质(D1-D3 格式)

        返回 True 表示加载成功.
        容错:若读取失败,保持默认值.
        """
        try:
            import openpyxl

            wb = openpyxl.load_workbook(filepath, data_only=True)
            ws = wb.worksheets[0]

            # 尝试 D1-D3 格式(沉砂池/初沉池旧格式)
            raw = [ws["D1"].value, ws["D2"].value, ws["D3"].value]
            if all(v is not None for v in raw):
                self.set_param("Q_design", float(raw[0]))
                self.set_param("Kz", float(raw[1]))
                self.set_param("Q_avg_daily", float(raw[0]) * 86.4 / float(raw[1]))
                # 简化:Q_avg = Q_max * 86.4 / Kz 仅是近似值
            return True
        except Exception as e:
            _log.warning("operation failed: %s", e, exc_info=True)
            return False

    def __repr__(self) -> str:
        return (
            f"<InputNode Q={self.get_param('Q_design'):.2f} m³/s "
            f"Kz={self.get_param('Kz'):.1f}>"
        )

"""
output_writer.py — 计算结果分类输出

两种输出模式:
  1. 管网输出: 写入原始 Excel 的「计算结果」sheet (沿用旧逻辑)
  2. 分类输出: 按构筑物分工作表输出到 output/ 目录
"""

import os
import sys
from datetime import datetime
from typing import Dict

import openpyxl
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side

# 维度标签解析
sys.path.insert(0, os.path.dirname(__file__))
from _logging import get_logger
from models.base import WaterQuality
from models.node_registry import is_io_node as _is_io_node
from ui.dimension_labels import format_dimension_row, split_dimensions

_log = get_logger(__name__)


def write_classified_output(
    results: Dict,
    executor,
    output_dir: str = "output",
    filename: str = None,
) -> str:
    """分类输出 — 按构筑物分 sheet 写入 Excel

    Args:
        results: GraphExecutor.execute() 返回的 {node_id: NodeResult}
        executor: GraphExecutor 实例
        output_dir: 输出目录
        filename: 文件名 (默认: shuichang_sheji_YYYYMMDD_HHMMSS.xlsx)

    Returns:
        输出文件路径
    """
    os.makedirs(output_dir, exist_ok=True)

    if filename is None:
        ts = datetime.now().strftime("%Y%m%d_%H%M%S")
        filename = f"shuichang_sheji_{ts}.xlsx"

    filepath = os.path.join(output_dir, filename)

    wb = openpyxl.Workbook()
    # 删除默认 sheet
    wb.remove(wb.active)

    # 样式
    hdr_font = Font(name="宋体", size=11, bold=True)
    title_font = Font(name="宋体", size=14, bold=True)
    nml_font = Font(name="宋体", size=10)
    thin_border = Border(
        left=Side("thin"),
        right=Side("thin"),
        top=Side("thin"),
        bottom=Side("thin"),
    )
    hdr_fill = PatternFill("solid", fgColor="D9E1F2")

    sheets_written = 0
    sheet_log = []  # 输出日志: [(name, status, detail)]

    for nid, r in results.items():
        if nid.startswith("_"):
            continue
        if not r.success:
            be = executor.get_node(nid)
            name = be.NODE_NAME if be else nid
            sheet_log.append(
                (
                    name,
                    be.NODE_CATEGORY if be else "",
                    "✗ 计算失败",
                    r.error_msg if hasattr(r, "error_msg") and r.error_msg else "",
                )
            )
            continue

        be = executor.get_node(nid)
        if not be:
            continue

        # 跳过输入/合并节点
        if _is_io_node(be.NODE_TYPE):
            continue

        name = be.NODE_NAME
        sheet_name = _safe_sheet_name(name)

        # 检查硬约束 — 失败时仍输出, 仅在标题标注 ⚠
        hard_failure_msg = ""
        try:
            from models.discretization import get_config

            hard_names = set(get_config(be.NODE_TYPE).get("constraint_names", []))
            hard_fails = [
                cn
                for cn, (passed, actual, limit, unit) in r.checks.items()
                if not passed and cn in hard_names
            ]
            if hard_fails:
                hard_failure_msg = f" ⚠ 硬约束未通过: {', '.join(hard_fails)}"
        except (KeyError, ImportError):
            pass

        ws = wb.create_sheet(title=sheet_name)

        # 标题
        ws.merge_cells("A1:D1")
        c = ws["A1"]
        c.value = f"{name} 设计计算结果{hard_failure_msg}"
        c.font = title_font
        c.alignment = Alignment(horizontal="center")
        ws.row_dimensions[1].height = 30

        row = 3

        # ── 一、原始设计参数 (用户选取的输入变量) ──
        if r.params:
            _write_section(ws, row, "原始设计参数", hdr_font, hdr_fill, thin_border)
            row += 1
            for ci, h in enumerate(["符号", "物理意义", "单位", "取值"], 1):
                c = ws.cell(row=row, column=ci, value=h)
                c.font = hdr_font
                c.fill = hdr_fill
                c.border = thin_border
                c.alignment = Alignment(horizontal="center")
            row += 1
            for k, v in r.params.items():
                sym, meaning, default_u = format_dimension_row(k, v, "", be.NODE_TYPE)
                val_str = f"{v:.2f}" if isinstance(v, float) else str(v)
                ws.cell(row=row, column=1, value=sym).font = nml_font
                ws.cell(row=row, column=2, value=meaning).font = nml_font
                ws.cell(row=row, column=3, value=default_u).font = nml_font
                ws.cell(row=row, column=4, value=val_str).font = nml_font
                for ci in range(1, 5):
                    ws.cell(row=row, column=ci).border = thin_border
                row += 1

        # ── 二、计算结果 (由计算得到的非尺寸变量) ──
        if r.dimensions:
            calc_dims, phys_dims, wq_in_out, wq_removal = split_dimensions(
                r.dimensions, r.dimension_categories
            )
            if calc_dims:
                row += 1
                _write_section(ws, row, "计算结果", hdr_font, hdr_fill, thin_border)
                row += 1
                for ci, h in enumerate(["符号", "物理意义", "单位", "取值"], 1):
                    c = ws.cell(row=row, column=ci, value=h)
                    c.font = hdr_font
                    c.fill = hdr_fill
                    c.border = thin_border
                    c.alignment = Alignment(horizontal="center")
                row += 1
                for k, v, u in calc_dims:
                    sym, meaning, default_u = format_dimension_row(k, v, u)
                    unit = u or default_u
                    val_str = f"{v:.2f}" if isinstance(v, float) else str(v)
                    ws.cell(row=row, column=1, value=sym).font = nml_font
                    ws.cell(row=row, column=2, value=meaning).font = nml_font
                    ws.cell(row=row, column=3, value=unit).font = nml_font
                    ws.cell(row=row, column=4, value=val_str).font = nml_font
                    for ci in range(1, 5):
                        ws.cell(row=row, column=ci).border = thin_border
                    row += 1

            # ── 三、构筑物尺寸 (仅物理几何参数) ──
            if phys_dims:
                row += 1
                _write_section(ws, row, "构筑物尺寸", hdr_font, hdr_fill, thin_border)
                row += 1
                for ci, h in enumerate(["符号", "物理意义", "单位", "取值"], 1):
                    c = ws.cell(row=row, column=ci, value=h)
                    c.font = hdr_font
                    c.fill = hdr_fill
                    c.border = thin_border
                    c.alignment = Alignment(horizontal="center")
                row += 1
                for k, v, u in phys_dims:
                    sym, meaning, default_u = format_dimension_row(k, v, u)
                    unit = u or default_u
                    val_str = f"{v:.2f}" if isinstance(v, float) else str(v)
                    ws.cell(row=row, column=1, value=sym).font = nml_font
                    ws.cell(row=row, column=2, value=meaning).font = nml_font
                    ws.cell(row=row, column=3, value=unit).font = nml_font
                    ws.cell(row=row, column=4, value=val_str).font = nml_font
                    for ci in range(1, 5):
                        ws.cell(row=row, column=ci).border = thin_border
                    row += 1

        # ── 四、水质处理效果 ──
        row += 1
        _write_section(ws, row, "水质处理效果", hdr_font, hdr_fill, thin_border)
        row += 1
        headers = [
            "指标",
            "进水(mg/L)",
            "出水(mg/L)",
            "去除率(%)",
            "排放标准(mg/L)",
            "达标",
        ]
        for ci, h in enumerate(headers, 1):
            c = ws.cell(row=row, column=ci, value=h)
            c.font = hdr_font
            c.fill = hdr_fill
            c.border = thin_border
        row += 1
        # 出水标准 — 按水类型自动选择
        if be.NODE_TYPE.startswith("kw_") or be.NODE_CATEGORY == "矿井水处理":
            effluent_std = {
                "BOD5": 4,
                "COD": 20,
                "SS": 70,
                "NH3N": 1.0,
                "TN": 1.0,
                "TP": 0.2,
            }
        else:
            effluent_std = {
                "BOD5": 10,
                "COD": 50,
                "SS": 10,
                "NH3N": 5,
                "TN": 15,
                "TP": 0.5,
            }
        wq_rows = 0
        for pk in ["BOD5", "COD", "SS", "NH3N", "TN", "TP"]:
            in_key = f"进水{pk}"
            out_key = f"出水{pk}"
            rm_key = f"{pk}去除率"
            # 优先用维度中的水质数据(标量计算产生)
            if in_key in r.dimensions and out_key in r.dimensions:
                inv = r.dimensions[in_key][0]
                outv = r.dimensions[out_key][0]
                rmv = r.dimensions.get(rm_key, ("—", ""))[0]
            # 回退: 从 inlet_quality / outlet_quality 提取(向量化缓存可能缺失)
            elif r.inlet_quality and r.outlet_quality:
                inv = getattr(r.inlet_quality, pk, None)
                outv = getattr(r.outlet_quality, pk, None)
                if inv is None or outv is None:
                    continue
                rmv = round((inv - outv) / inv * 100, 1) if inv > 0 else 0
            # 回退: 从 removal_rates 估算
            elif r.removal_rates and r.removal_rates.get(pk, 0) > 0:
                inv = getattr(WaterQuality(), pk, 200 if pk == "BOD5" else 400)
                rate = r.removal_rates.get(pk, 0)
                outv = round(inv * (1 - rate), 2)
                rmv = round(rate * 100, 1)
            else:
                continue
            eff = effluent_std.get(pk, 0)
            ok_val = "✓" if outv <= eff else "✗"
            ws.cell(row=row, column=1, value=pk).font = nml_font
            ws.cell(row=row, column=2, value=round(inv, 2)).font = nml_font
            ws.cell(row=row, column=3, value=round(outv, 2)).font = nml_font
            ws.cell(row=row, column=4, value=f"{rmv}%").font = nml_font
            ws.cell(row=row, column=5, value=eff).font = nml_font
            ws.cell(row=row, column=6, value=ok_val).font = nml_font
            for ci in range(1, 7):
                ws.cell(row=row, column=ci).border = thin_border
            row += 1
            wq_rows += 1
        if wq_rows == 0:
            ws.cell(row=row, column=1, value="(无水质数据, 请按 F5 计算)").font = (
                nml_font
            )
            for ci in range(1, 7):
                ws.cell(row=row, column=ci).border = thin_border
            row += 1
            wq_rows += 1

        # ── 五、约束校核 ──
        if r.checks:
            row += 1
            _write_section(ws, row, "约束校核", hdr_font, hdr_fill, thin_border)
            row += 1
            for cn, (passed, actual, limit, unit) in r.checks.items():
                ws.cell(row=row, column=1, value=cn).font = nml_font
                ws.cell(row=row, column=2, value=f"{actual} {unit}").font = nml_font
                ws.cell(row=row, column=3, value=limit).font = nml_font
                ws.cell(row=row, column=4, value="✓" if passed else "✗").font = nml_font
                for ci in range(1, 5):
                    ws.cell(row=row, column=ci).border = thin_border
                row += 1

        # ── 安全系数 ──
        if hasattr(r, "robustness") and r.robustness > 0:
            row += 1
            _write_section(ws, row, "安全评估", hdr_font, hdr_fill, thin_border)
            row += 1
            ws.cell(row=row, column=1, value="综合安全系数").font = nml_font
            c = ws.cell(row=row, column=2, value=round(r.robustness, 4))
            c.font = Font(
                name="宋体",
                size=12,
                bold=True,
                color=(
                    "00AA00"
                    if r.robustness > 0.2
                    else "FF8800" if r.robustness > 0.05 else "FF0000"
                ),
            )
            ws.cell(row=row, column=3, value="0~1 (越大越安全)").font = nml_font
            for ci in range(1, 4):
                ws.cell(row=row, column=ci).border = thin_border
            row += 1

        # 列宽
        ws.column_dimensions["A"].width = 20
        ws.column_dimensions["B"].width = 15
        ws.column_dimensions["C"].width = 15
        ws.column_dimensions["D"].width = 12
        ws.column_dimensions["E"].width = 15
        ws.column_dimensions["F"].width = 10

        sheets_written += 1

        # ── 单构筑物高程 sheet ──
        elev = getattr(r, "elevation", None)
        if elev is not None:
            elev_name = _safe_sheet_name(name + "_高程")
            ws_elev = wb.create_sheet(title=elev_name)
            _write_single_elevation_sheet(
                ws_elev,
                name,
                elev,
                be.NODE_TYPE,
                r,
                hdr_font,
                hdr_fill,
                title_font,
                nml_font,
                thin_border,
            )
            sheets_written += 1

        status = "⚠ 硬约束" if hard_failure_msg else "✓"
        sheet_log.append(
            (
                name,
                be.NODE_CATEGORY,
                status,
                hard_failure_msg.replace(" ⚠ 硬约束未通过: ", ""),
            )
        )

    # 记录跳过的失败节点
    for nid, r in results.items():
        if nid.startswith("_"):
            continue
        be = executor.get_node(nid)
        if not be or r.success:
            continue
        if _is_io_node(be.NODE_TYPE):
            continue
        sheet_log.append((be.NODE_NAME, be.NODE_CATEGORY, "✗ 计算失败", ""))

    # ── 高程计算结果 sheet ──
    _write_elevation_sheet(
        wb, results, executor, hdr_font, hdr_fill, title_font, nml_font, thin_border
    )

    # ── 汇总 sheet ──
    if sheets_written > 0 or sheet_log:
        ws_sum = wb.create_sheet(title="汇总", index=0)
        ws_sum.merge_cells("A1:D1")
        ws_sum["A1"].value = "污水处理厂设计计算汇总"
        ws_sum["A1"].font = title_font
        ws_sum["A1"].alignment = Alignment(horizontal="center")
        ws_sum.row_dimensions[1].height = 30

        row = 3
        _write_section(ws_sum, row, "各构筑物概况", hdr_font, hdr_fill, thin_border)
        row += 1
        for ci, h in enumerate(["构筑物", "类型", "关键尺寸", "状态"], 1):
            c = ws_sum.cell(row=row, column=ci, value=h)
            c.font = hdr_font
            c.fill = hdr_fill
            c.border = thin_border
        row += 1
        for nid, r in results.items():
            if nid.startswith("_"):
                continue
            be = executor.get_node(nid)
            if not be or not r.success:
                continue
            if _is_io_node(be.NODE_TYPE):
                continue
            dim_summary = _dim_summary(r.dimensions)
            ws_sum.cell(row=row, column=1, value=be.NODE_NAME).font = nml_font
            ws_sum.cell(row=row, column=2, value=be.NODE_CATEGORY).font = nml_font
            ws_sum.cell(row=row, column=3, value=dim_summary).font = nml_font
            ws_sum.cell(row=row, column=4, value="✓").font = nml_font
            for ci in range(1, 5):
                ws_sum.cell(row=row, column=ci).border = thin_border
            row += 1

        # ── 输出日志 ──
        row += 1
        _write_section(ws_sum, row, "输出日志", hdr_font, hdr_fill, thin_border)
        row += 1
        for ci, h in enumerate(["构筑物", "类型", "状态", "备注"], 1):
            c = ws_sum.cell(row=row, column=ci, value=h)
            c.font = hdr_font
            c.fill = hdr_fill
            c.border = thin_border
        row += 1
        for name, cat, status, detail in sheet_log:
            ws_sum.cell(row=row, column=1, value=name).font = nml_font
            ws_sum.cell(row=row, column=2, value=cat).font = nml_font
            ws_sum.cell(row=row, column=3, value=status).font = nml_font
            ws_sum.cell(row=row, column=4, value=detail).font = nml_font
            for ci in range(1, 5):
                ws_sum.cell(row=row, column=ci).border = thin_border
            row += 1

    wb.save(filepath)
    return filepath


def _safe_sheet_name(name: str) -> str:
    """生成安全的 sheet 名称 (≤31字符, 不含特殊字符)"""
    safe = name.replace("/", "").replace("\\", "").replace("*", "").replace("?", "")
    safe = safe.replace("[", "").replace("]", "").replace(":", "")
    return safe[:31]

    # ═══════════════ 事件回调 ═══════════════


def _write_single_elevation_sheet(
    ws,
    name,
    elev,
    node_type,
    result,
    hdr_font,
    hdr_fill,
    title_font,
    nml_font,
    thin_border,
):
    """写入单个构筑物的高程计算表"""
    is_pipe = node_type == "gdys_stss"
    depth_label = "管内水深" if is_pipe else "有效水深"
    bottom_label = "管内底标高" if is_pipe else "池底标高"

    ws.merge_cells("A1:D1")
    c = ws["A1"]
    c.value = f"{name} — 高程计算结果"
    c.font = title_font
    c.alignment = Alignment(horizontal="center")
    ws.row_dimensions[1].height = 30

    headers = ["符号", "物理意义", "单位", "取值"]
    widths = [10, 32, 8, 14]
    row = 3
    for ci, (h, w) in enumerate(zip(headers, widths), 1):
        c = ws.cell(row=row, column=ci, value=h)
        c.font = hdr_font
        c.fill = hdr_fill
        c.border = thin_border
        c.alignment = Alignment(horizontal="center")
        ws.column_dimensions[chr(64 + ci)].width = w
    row += 1

    # 从 result.dimensions 提取分项水头损失
    h_f, h_m = 0.0, 0.0
    if result and result.dimensions:
        h_f = result.dimensions.get("沿程水头损失 h_f", (0.0, ""))[0]
        h_m = result.dimensions.get("局部水头损失 h_m", (0.0, ""))[0]

    rows_data = [
        ("输入条件", None, None, None),
        ("Z_up", "上游水面标高", "m", elev.upstream_water_elevation),
        ("Z_ground", "地面标高", "m", elev.ground_elevation),
        ("水头损失", None, None, None),
        ("Δh", "总水头损失", "m", elev.head_loss),
        ("", f"  ↳ {elev.head_loss_detail[:100]}", "", ""),
    ]
    # 有分项水头损失时展开显示
    if h_f > 0 or h_m > 0:
        rows_data.append(("h_f", "  沿程水头损失 (Manning公式)", "m", h_f))
        rows_data.append(("h_m", "  局部水头损失 (ξ·v²/2g)", "m", h_m))

    rows_data += [
        ("高程计算", None, None, None),
        ("Z_water", "本节点水面标高", "m", elev.water_elevation),
        ("h_eff" if not is_pipe else "DN", depth_label, "m", elev.effective_depth),
        ("Z_bottom", bottom_label, "m", elev.bottom_elevation),
    ]
    if not is_pipe:
        rows_data += [
            ("h_super", "超高", "m", elev.super_elevation),
            (
                "h_cover",
                "埋深 = 地面 - 池底",
                "m",
                elev.ground_elevation - elev.bottom_elevation,
            ),
        ]

    for item in rows_data:
        if item[1] is None:
            ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=4)
            c = ws.cell(row=row, column=1, value=f"── {item[0]} ──")
            c.font = hdr_font
            c.fill = hdr_fill
            c.border = thin_border
        else:
            for ci, val in enumerate(item, 1):
                v = round(val, 3) if isinstance(val, float) else (val or "")
                c = ws.cell(row=row, column=ci, value=v)
                c.font = nml_font
                c.border = thin_border
        row += 1


def _write_elevation_sheet(
    wb, results, executor, hdr_font, hdr_fill, title_font, nml_font, thin_border
):
    """写入高程计算结果 sheet"""
    # 收集高程数据
    elevation_rows = []
    try:
        order = executor.topological_order()
    except Exception as e:
        _log.warning("operation failed: %s", e, exc_info=True)
        order = list(results.keys())

    for nid in order:
        r = results.get(nid)
        if not r or not r.success:
            continue
        be = executor.get_node(nid)
        if not be:
            continue
        elev = getattr(r, "elevation", None)
        if elev is None:
            continue
        elevation_rows.append((be.NODE_NAME, elev))

    if not elevation_rows:
        return

    ws = wb.create_sheet(title="高程计算结果")
    ws.merge_cells("A1:H1")
    c = ws["A1"]
    c.value = "污水处理厂高程计算结果"
    c.font = title_font
    c.alignment = Alignment(horizontal="center")
    ws.row_dimensions[1].height = 30

    row = 3
    headers = [
        "构筑物",
        "地面标高(m)",
        "池底标高(m)",
        "水面标高(m)",
        "有效水深(m)",
        "超高(m)",
        "水头损失(m)",
        "计算公式",
    ]
    for ci, h in enumerate(headers, 1):
        c = ws.cell(row=row, column=ci, value=h)
        c.font = hdr_font
        c.fill = hdr_fill
        c.border = thin_border
        c.alignment = Alignment(horizontal="center")
    row += 1

    for name, elev in elevation_rows:
        ws.cell(row=row, column=1, value=name).font = nml_font
        ws.cell(row=row, column=2, value=round(elev.ground_elevation, 3)).font = (
            nml_font
        )
        ws.cell(row=row, column=3, value=round(elev.bottom_elevation, 3)).font = (
            nml_font
        )
        ws.cell(row=row, column=4, value=round(elev.water_elevation, 3)).font = nml_font
        ws.cell(row=row, column=5, value=round(elev.effective_depth, 3)).font = nml_font
        ws.cell(row=row, column=6, value=round(elev.super_elevation, 3)).font = nml_font
        ws.cell(row=row, column=7, value=round(elev.head_loss, 3)).font = nml_font
        ws.cell(row=row, column=8, value=elev.head_loss_detail[:60]).font = nml_font
        for ci in range(1, 9):
            ws.cell(row=row, column=ci).border = thin_border
        row += 1

    # 列宽
    ws.column_dimensions["A"].width = 16
    ws.column_dimensions["B"].width = 12
    ws.column_dimensions["C"].width = 12
    ws.column_dimensions["D"].width = 12
    ws.column_dimensions["E"].width = 12
    ws.column_dimensions["F"].width = 10
    ws.column_dimensions["G"].width = 12
    ws.column_dimensions["H"].width = 30


def _is_physical_dimension(name: str) -> bool:
    """判断维度名是否属于构筑物物理尺寸 (非计算值/运行参数)

    物理尺寸 = 施工人员建造构筑物所需的几何参数:
      数量、长度、宽度、高度、直径、面积、容积、比例
    """
    phys_keywords = [
        # 数量
        "池数",
        "格数",
        "系列数",
        "渠道数",
        "格栅台数",
        "滤池格数",
        "磁盘台数",
        "砂斗个数",
        "磁盘片数",
        "栅条间隙数",
        "滤头数量",
        "灯管排数",
        "灯管总数",
        # 线性尺寸
        "池径",
        "池长",
        "池宽",
        "直径",
        "长度",
        "宽度",
        "总高度",
        "有效水深",
        "超高",
        "池底坡降",
        "泥斗高度",
        "锥体高度",
        "圆柱段高度",
        "砂斗深度",
        "中心管径",
        "渠宽",
        "栅后总高",
        "滤池总高度",
        "进水渠宽",
        "栅槽宽度",
        "单格宽度",
        "砂斗上口直径",
        "滗水高度",
        "污泥层高度",
        "安全距离",
        "浓缩区高度",
        "磁盘直径",
        "磁盘间隙",
        "灯管长度",
        "灯管间隙",
        "排沙口直径",
        # 面积
        "沉淀面积",
        "总过滤面积",
        "单格面积",
        "调节池总面积",
        "沉砂池总面积",
        "总面积",
        "单格过滤面积",
        "出水堰长",
        "磁盘组长度",
        "渠道总长",
        "栅槽总长",
        "单格长度",
        "沉淀区面积",
        "进水渠断面积",
        # 容积
        "有效容积",
        "单池有效容积",
        "总有效容积",
        "主反应区总容积",
        "单池主反应区容积",
        "选择区容积",
        "单池总有效容积",
        "混合区容积",
        "絮凝区容积",
        "单系列总容积",
        "污泥区总容积",
        "砂斗总容积",
        "锥体实际容积",
        "圆柱储砂容积",
        "单池需贮泥容积",
        "2日贮泥容积",
        # 比例 / 坡度
        "径深比",
        "长宽比",
        "宽高比",
        "充水比",
        "池底坡度",
    ]
    for kw in phys_keywords:
        if kw in name:
            return True
    return False

    # ═══════════════ 写入/输出 ═══════════════


def _write_section(ws, row, title, hdr_font, hdr_fill, border):
    """写入章节标题 — 先解除旧合并再合并, 避免 MergedCell 读写冲突.

    仅对合并区域的左上角单元格 (row, 1) 设置样式.
    Excel 合并单元格的边框由左上角单元格决定,
    避免访问 MergedCell 对象 (openpyxl 3.1.5 中可能触发 value 只读错误).
    """
    try:
        ws.unmerge_cells(start_row=row, start_column=1, end_row=row, end_column=6)
    except ValueError:
        pass  # 单元格未合并,无需解除
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=6)
    c = ws.cell(row=row, column=1)
    c.value = f"▎{title}"
    c.font = hdr_font
    c.fill = hdr_fill
    c.border = border


def _dim_summary(dims: Dict) -> str:
    """生成尺寸摘要"""
    key_dims = [
        "池数",
        "池径 D",
        "池长 L",
        "池宽 B",
        "总高度 H",
        "有效水深 h2",
        "格栅台数",
        "滤池格数",
        "渠道数",
    ]
    parts = []
    for k in key_dims:
        if k in dims:
            v, u = dims[k]
            if isinstance(v, float):
                parts.append(f"{k}={v:.2f}{u}")
            else:
                parts.append(f"{k}={v}{u}")
    return ", ".join(parts[:4]) if parts else "—"

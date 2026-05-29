"""
solution_browser.py — 方案浏览器 UI 组件

替代原有的滑块面板,提供:
  - 参数选择器(Combobox,每参数一个)
  - 可行方案表格(Treeview,按成本排序)
  - 方案详情面板
  - 应用方案按钮
"""

from __future__ import annotations

import os
import sys
import tkinter as tk
from tkinter import ttk
from typing import Callable, Dict, List, Optional

import numpy as np

sys.path.insert(0, os.path.join(os.path.dirname(__file__), ".."))

from _logging import get_logger
from models.base import WaterFlow, WaterQuality
from models.discretization import get_config, get_free_keys
from models.solution_space import Solution, get_engine

from .dimension_labels import format_dimension_row, format_param_value

_log = get_logger(__name__)


class SolutionBrowser(tk.Frame):
    """方案浏览器 — 替换参数滑块面板"""

    def __init__(
        self,
        parent,
        on_apply: Optional[Callable] = None,
        bg: str = "#252525",
        **kw,
    ):
        super().__init__(parent, bg=bg, **kw)
        self._bg = bg
        self._apply_callback = on_apply  # 外部回调 (不与 _on_apply 方法重名)
        self._solutions: List[Solution] = []
        self._selected_idx: int = -1
        self._backend_node = None
        self._node_type: str = ""
        self._flow: Optional[WaterFlow] = None
        self._quality: Optional[WaterQuality] = None

        # 参数筛选器状态 {key: current_value or None (全部)}
        self._filters: Dict[str, Optional[float]] = {}

        # ── 构建 UI ──
        self._build_ui()

    # ═══════════════════ 公开 API ═══════════════════

    def load_node(
        self,
        backend_node,
        flow: WaterFlow,
        quality: WaterQuality,
        force_recompute: bool = False,
    ) -> None:
        """为指定节点加载方案空间"""
        self._backend_node = backend_node
        self._node_type = backend_node.NODE_TYPE
        self._flow = flow
        self._quality = quality
        self._sludge = None  # 清除污泥上下文，避免状态泄漏
        self._solutions = []
        self._selected_idx = -1
        self._filters = {}

        # 隐藏内容,显示加载提示
        self._clear_all()
        self._status_label.config(text="正在计算可行方案...")
        self.update_idletasks()

        try:
            engine = get_engine()
            solutions = engine.enumerate(
                self._node_type, flow, quality, force_recompute=force_recompute
            )
            self._solutions = solutions
            if solutions:
                self._status_label.config(text=f"共 {len(solutions)} 个可行方案")
            else:
                self._status_label.config(text="无可行方案")
                # ── v5.4-s7: 无可行解时在面板上显示增强诊断 ──
                self._show_no_solution_hint(engine, flow, quality)
                return  # ← v5.4-s7 fix: 跳过后续 _build_filter_ui,
                # 否则 filter UI 会 destroy 诊断提示组件
        except NotImplementedError:
            self._status_label.config(text="该模块暂不支持方案浏览")
            self._build_filter_ui([])
            return
        except Exception as e:
            self._status_label.config(text=f"计算失败: {e}")
            return

        # 构建筛选器 UI
        self._build_filter_ui(self._solutions)
        self._refresh_table()
        # ── 自动选中与当前节点参数匹配的方案 ──
        self._select_applied()

    def load_sludge_node(self, backend_node, sludge) -> None:
        """为污泥处理节点加载方案空间 (使用 SludgeFlow 上下文)"""
        self._backend_node = backend_node
        self._node_type = backend_node.NODE_TYPE
        self._sludge = sludge
        self._flow = None     # 清除水量上下文，避免状态泄漏到标量验证
        self._quality = None  # 清除水质上下文
        self._solutions = []
        self._selected_idx = -1
        self._filters = {}

        self._clear_all()
        self._status_label.config(text="正在计算污泥可行方案...")
        self.update_idletasks()

        try:
            from models.solution_space import enumerate_sludge_solutions

            solutions = enumerate_sludge_solutions(self._node_type, sludge)
            self._solutions = solutions
            self._status_label.config(
                text=f"共 {len(solutions)} 个可行方案" if solutions else "无可行方案"
            )
        except NotImplementedError:
            self._status_label.config(text="该模块暂不支持方案浏览")
            self._build_filter_ui([])
            return
        except Exception as e:
            self._status_label.config(text=f"计算失败: {e}")
            return

        # 构建筛选器 UI
        self._build_filter_ui(self._solutions)
        self._refresh_table()
        self._select_applied()

    # ═══════════════ 查询/获取 ═══════════════
    def get_selected_params(self) -> Dict[str, float]:
        """获取当前选中方案的参数"""
        if 0 <= self._selected_idx < len(self._solutions):
            return dict(self._solutions[self._selected_idx].params)
        return {}

    # ═══════════════════ UI 构建 ═══════════════════

    def _build_ui(self):
        """初始化 UI 框架"""
        bg = self._bg

        # ── 标题行 ──
        header = tk.Frame(self, bg=bg)
        header.pack(fill=tk.X, padx=8, pady=(8, 4))
        self._title_label = tk.Label(
            header,
            text="▎方案浏览器",
            bg=bg,
            fg="#fff",
            font=("Microsoft YaHei", 12, "bold"),
        )
        self._title_label.pack(side=tk.LEFT)

        # 导出按钮
        self._export_btn = tk.Button(
            header,
            text="📊 导出全部方案",
            bg="#3a6b3a",
            fg="#fff",
            activebackground="#4a8b4a",
            activeforeground="#fff",
            font=("Microsoft YaHei", 8),
            relief=tk.FLAT,
            padx=8,
            pady=2,
            command=self._export_to_excel,
        )
        self._export_btn.pack(side=tk.RIGHT, padx=4)

        # ── 状态标签 ──
        self._status_label = tk.Label(
            self,
            text="",
            bg=bg,
            fg="#888",
            font=("Microsoft YaHei", 8),
        )
        self._status_label.pack(fill=tk.X, padx=8)

        # ── 筛选器容器 ──
        self._filter_frame = tk.Frame(self, bg=bg)
        self._filter_frame.pack(fill=tk.X, padx=6, pady=2)

        # ── 表格 ──
        table_frame = tk.Frame(self, bg=bg)
        table_frame.pack(fill=tk.BOTH, expand=True, padx=6, pady=2)

        self._tree = ttk.Treeview(
            table_frame,
            columns=("rank", "cost", "summary"),
            show="headings",
            height=8,
        )
        self._tree.heading("rank", text="#")
        self._tree.heading("cost", text="成本(万元)")
        self._tree.heading("summary", text="关键尺寸")
        self._tree.column("rank", width=30, anchor="center")
        self._tree.column("cost", width=80, anchor="e")
        self._tree.column("summary", width=280, anchor="w")

        scrollbar = ttk.Scrollbar(
            table_frame, orient=tk.VERTICAL, command=self._tree.yview
        )
        self._tree.configure(yscrollcommand=scrollbar.set)
        self._tree.pack(side=tk.LEFT, fill=tk.BOTH, expand=True)
        scrollbar.pack(side=tk.RIGHT, fill=tk.Y)

        self._tree.bind("<<TreeviewSelect>>", self._on_row_selected)

        # ── 详情面板 (4列 Treeview: 符号|物理意义|单位|取值, 固定高度不抢占按钮空间) ──
        detail_frame = tk.Frame(self, bg=bg)
        detail_frame.pack(fill=tk.X, padx=6, pady=2)  # 不 expand, 固定高度

        self._detail_tree = ttk.Treeview(
            detail_frame,
            columns=("symbol", "meaning", "unit", "value"),
            show="headings",
            height=4,
        )
        self._detail_tree.heading("symbol", text="符号")
        self._detail_tree.heading("meaning", text="物理意义")
        self._detail_tree.heading("unit", text="单位")
        self._detail_tree.heading("value", text="取值")
        self._detail_tree.column("symbol", width=70, anchor="center")
        self._detail_tree.column("meaning", width=140, anchor="w")
        self._detail_tree.column("unit", width=55, anchor="center")
        self._detail_tree.column("value", width=110, anchor="e")
        style = ttk.Style()
        style.configure(
            "Treeview",
            background="#1a1a1a",
            foreground="#ccc",
            fieldbackground="#1a1a1a",
            rowheight=20,
        )
        style.configure("Treeview.Heading", background="#2d2d2d", foreground="#ffaa44")
        style.map("Treeview", background=[("selected", "#3a5a1a")])
        detail_scroll = ttk.Scrollbar(
            detail_frame, orient=tk.VERTICAL, command=self._detail_tree.yview
        )
        self._detail_tree.configure(yscrollcommand=detail_scroll.set)
        self._detail_tree.pack(side=tk.LEFT, fill=tk.BOTH, expand=True)
        detail_scroll.pack(side=tk.RIGHT, fill=tk.Y)

        # ── 约束校核 (底部, 最小高度) ──
        self._detail_check_text = tk.Text(
            self,
            bg="#1a1a1a",
            fg="#ccc",
            font=("Consolas", 8),
            wrap=tk.WORD,
            relief=tk.FLAT,
            borderwidth=0,
            height=2,
        )
        self._detail_check_text.pack(fill=tk.X, padx=6, pady=2)

        # ── 按钮行 ──
        btn_frame = tk.Frame(self, bg=bg)
        btn_frame.pack(fill=tk.X, padx=6, pady=4)
        ttk.Button(btn_frame, text="▶ 应用选中方案", command=self._on_apply).pack(
            side=tk.LEFT,
            padx=2,
        )
        ttk.Button(btn_frame, text="★ 推荐方案", command=self._select_recommended).pack(
            side=tk.LEFT,
            padx=2,
        )

        # ── 排序切换 ──
        self._sort_mode = "cost"  # "cost" | "robustness"
        self._sort_btn = ttk.Button(
            btn_frame,
            text="💰 成本排序",
            command=self._on_toggle_sort,
            width=12,
        )
        self._sort_btn.pack(side=tk.RIGHT, padx=2)

        ttk.Button(btn_frame, text="⟲ 刷新", command=self._on_refresh).pack(
            side=tk.RIGHT,
            padx=2,
        )

    # ═══════════════ 无可行解诊断 (v5.4-s7 增强) ═══════════════

    def _show_no_solution_hint(self, engine, flow, quality):
        """当无可行方案时, 在面板上显示增强诊断 (v5.4-s7)

        包含三个模块:
          1. 最小冲突集 — 无法同时满足的约束子集
          2. 每约束提示 — 具体调整哪个参数的哪个方向
          3. 通过率条 — 可视化每约束的通过比例
        """
        from models.discretization import get_config

        cfg = get_config(self._node_type)
        constraint_keys = cfg.get("constraint_keys", [])
        constraint_names = cfg.get("constraint_names", [])
        free_keys = list(cfg.get("free", {}).keys())

        if not constraint_keys:
            self._status_label.config(text="无可行方案 (无约束可诊断)")
            return

        # 用全量 free 值枚举一次以获得结果数组进行诊断
        free_vals = [cfg["free"][k] for k in free_keys]
        meshes = np.meshgrid(*free_vals, indexing="ij")
        grid = {k: mesh.ravel() for k, mesh in zip(free_keys, meshes)}
        fixed = cfg.get("fixed", {})
        results = engine._compute_vectorized(
            self._node_type, grid, flow, quality, fixed
        )
        engine._filter_feasible(
            results,
            constraint_keys,
            constraint_names,
            cfg.get("constraint_limits"),
            self._node_type,
        )
        diag = engine._diagnose_infeasibility(
            results, constraint_keys, constraint_names, cfg
        )

        # ── 清除旧内容 ──
        for w in self._filter_frame.winfo_children():
            w.destroy()

        hint = tk.Frame(self._filter_frame, bg=self._bg)
        hint.pack(fill=tk.X, pady=8, padx=10)

        # ── 标题 ──
        tk.Label(
            hint,
            text="⚠ 无可行方案 — 诊断报告",
            bg=self._bg,
            fg="#ff6644",
            font=("Microsoft YaHei", 11, "bold"),
        ).pack(anchor="w")

        # ── 1. 最小冲突集 ──
        conflict_set = diag.get("conflict_set", [])
        if conflict_set:
            cf_frame = tk.Frame(hint, bg="#3a1a1a", bd=1, relief=tk.GROOVE)
            cf_frame.pack(fill=tk.X, pady=(6, 4), ipady=4)
            tk.Label(
                cf_frame,
                text=f"⚡ 最小冲突集: 以下 {len(conflict_set)} 个约束无法同时满足",
                bg="#3a1a1a",
                fg="#ff9966",
                font=("Microsoft YaHei", 9, "bold"),
                wraplength=360,
            ).pack(anchor="w", padx=8, pady=(4, 0))
            tk.Label(
                cf_frame,
                text="放宽其中任一约束即可使方案可行",
                bg="#3a1a1a",
                fg="#cc8866",
                font=("Microsoft YaHei", 8),
                wraplength=360,
            ).pack(anchor="w", padx=8)
            for cname in conflict_set:
                hint_text = diag.get("hints", {}).get(cname, "")
                line = f"  • {cname}"
                if hint_text:
                    line += f" → {hint_text}"
                tk.Label(
                    cf_frame,
                    text=line,
                    bg="#3a1a1a",
                    fg="#ff8866",
                    font=("Microsoft YaHei", 8),
                    wraplength=360,
                    justify="left",
                ).pack(anchor="w", padx=12)

        # ── 2. 逐约束提示 + 通过率 ──
        constraint_rates = diag.get("constraint_rates", [])
        if constraint_rates:
            detail = tk.Frame(hint, bg=self._bg)
            detail.pack(fill=tk.X, pady=(4, 0))
            tk.Label(
                detail,
                text="▎逐约束分析",
                bg=self._bg,
                fg="#ffaa44",
                font=("Microsoft YaHei", 9, "bold"),
            ).pack(anchor="w")

            for cname, passed, total in constraint_rates:
                ratio = passed / total if total > 0 else 0
                color = "#55cc55" if ratio >= 1.0 else "#cc9955" if ratio > 0.5 else "#cc5555"
                bar_len = int(ratio * 15)
                bar = "█" * bar_len + "░" * (15 - bar_len)

                row = tk.Frame(detail, bg=self._bg)
                row.pack(fill=tk.X, pady=1)

                tk.Label(
                    row,
                    text=f"  {cname}",
                    bg=self._bg,
                    fg="#888",
                    font=("Microsoft YaHei", 8),
                    width=24,
                    anchor="w",
                ).pack(side=tk.LEFT)
                tk.Label(
                    row,
                    text=f"{bar}",
                    bg=self._bg,
                    fg=color,
                    font=("Consolas", 7),
                ).pack(side=tk.LEFT)
                tk.Label(
                    row,
                    text=f" {ratio:.0%}",
                    bg=self._bg,
                    fg=color,
                    font=("Consolas", 8, "bold"),
                    width=5,
                    anchor="w",
                ).pack(side=tk.LEFT)

                hint_text = diag.get("hints", {}).get(cname, "")
                if hint_text and ratio < 1.0:
                    tk.Label(
                        row,
                        text=f"→ {hint_text}",
                        bg=self._bg,
                        fg="#aaa",
                        font=("Microsoft YaHei", 8),
                        wraplength=200,
                        justify="left",
                    ).pack(side=tk.LEFT)

    # ═══════════════ 筛选器 UI ═══════════════

    def _build_filter_ui(self, solutions: List[Solution]) -> None:
        """为当前模块构建参数筛选 Combobox"""
        # 清除旧筛选器
        for w in self._filter_frame.winfo_children():
            w.destroy()

        if not solutions or not self._node_type:
            return

        cfg = get_config(self._node_type)
        free_keys = get_free_keys(self._node_type)
        free_vals = cfg["free"]

        # 为每个自由变量创建一行: 标签 + Combobox
        for i, key in enumerate(free_keys):
            row = tk.Frame(self._filter_frame, bg=self._bg)
            row.pack(fill=tk.X, pady=1)

            tk.Label(
                row,
                text=key,
                bg=self._bg,
                fg="#aaa",
                font=("Consolas", 8),
                width=10,
                anchor="e",
            ).pack(side=tk.LEFT, padx=(0, 4))

            values = ["全部"] + [str(v) for v in free_vals[key]]
            cb = ttk.Combobox(
                row,
                values=values,
                state="readonly",
                width=12,
            )
            cb.set("全部")
            cb.pack(side=tk.LEFT)

            # 绑定筛选事件
            cb.bind(
                "<<ComboboxSelected>>",
                lambda e, k=key, c=cb: self._on_filter_changed(k, c.get()),
            )

    # ═══════════════════ 筛选 + 表格 ═══════════════════

    def _on_filter_changed(self, key: str, value_str: str):
        """筛选条件变更"""
        if value_str == "全部":
            self._filters[key] = None
        else:
            self._filters[key] = float(value_str)
        self._refresh_table()
        self._select_applied()

    def _on_toggle_sort(self):
        """切换排序模式"""
        if self._sort_mode == "cost":
            self._sort_mode = "robustness"
            self._sort_btn.config(text="📐 安全排序")
        else:
            self._sort_mode = "cost"
            self._sort_btn.config(text="💰 成本排序")
        self._refresh_table()
        self._select_applied()

    # ═══════════════ 面板刷新 ═══════════════
    def _refresh_table(self):
        """根据当前筛选条件刷新表格并按当前排序模式排列"""
        for item in self._tree.get_children():
            self._tree.delete(item)
        if not self._solutions:
            return

        # 筛选
        filtered = self._solutions
        for key, val in self._filters.items():
            if val is not None:
                filtered = [
                    s for s in filtered if abs(s.params.get(key, 0) - val) < 1e-9
                ]

        # 排序
        if self._sort_mode == "cost":
            filtered = sorted(
                filtered,
                key=lambda s: s.cost_wan_yuan if s.cost_wan_yuan > 0 else float("inf"),
            )
        else:
            filtered = sorted(filtered, key=lambda s: -s.robustness)  # 裕度大的在前

        # 填充表格
        applied_params = self._get_applied_params()
        for i, sol in enumerate(filtered):
            summary = self._format_summary(sol)
            cost_str = f"{sol.cost_wan_yuan:.1f}" if sol.cost_wan_yuan > 0 else "—"
            if self._sort_mode == "robustness":
                cost_str += f" [{sol.robustness:.2f}]"
            # 标记已应用方案 (绿色高亮)
            tags = []
            if i == 0 and self._sort_mode == "cost":
                tags.append("recommended")
            if self._is_solution_applied(sol, applied_params):
                tags.append("applied")
                cost_str = "✓ " + cost_str
            tag = " ".join(tags) if tags else ""
            self._tree.insert(
                "",
                tk.END,
                iid=str(i),
                values=(i + 1, cost_str, summary),
                tags=(tag,),
            )

        self._tree.tag_configure("recommended", background="#2a3a1a")
        self._tree.tag_configure(
            "applied", foreground="#55ff55", font=("Consolas", 9, "bold")
        )
        self._status_label.config(
            text=f"显示 {len(filtered)}/{len(self._solutions)} 个方案"
            + f" (按{'安全裕度' if self._sort_mode == 'robustness' else '成本'}排序)"
        )

    # ═══════════════ 格式化 ═══════════════
    def _format_summary(self, sol: Solution) -> str:
        """生成关键尺寸摘要文本"""
        dims = sol.dimensions
        parts = []

        # 常见尺寸字段(按优先级)
        key_dims = [
            "L",
            "B",
            "D",
            "H_total",
            "V_total",
            "L_total",
            "B_channel",
            "V_eff_actual",
            "A_settle",
            "A_single",
        ]
        for key in key_dims:
            if key in dims:
                val, unit = (
                    dims[key] if isinstance(dims[key], tuple) else (dims[key], "")
                )
                if isinstance(val, float):
                    parts.append(f"{key}={val:.1f}{unit}")
                else:
                    parts.append(f"{key}={val}")

        return "  ".join(parts[:4])  # 最多显示4项

    # ═══════════════ 事件回调 ═══════════════
    def _on_row_selected(self, event):
        """表格行选中 → 更新详情面板"""
        selection = self._tree.selection()
        if not selection:
            self._clear_detail()
            return

        idx = int(selection[0])
        filtered = self._get_filtered_solutions()
        if 0 <= idx < len(filtered):
            sol = filtered[idx]
            self._selected_idx = self._solutions.index(sol)
            self._show_detail(sol)

    # ═══════════════ 面板显示 ═══════════════
    def _show_detail(self, sol: Solution):
        """显示方案详细信息 (4列 Treeview)"""
        tree = self._detail_tree
        for item in tree.get_children():
            tree.delete(item)
        self._detail_check_text.delete("1.0", tk.END)

        # ── 参数 ──
        sec_p = tree.insert(
            "",
            tk.END,
            values=("", "── 设计参数 ──", "", ""),
            tags=("section",),
            open=True,
        )
        for k, v in sol.params.items():
            sym, meaning, _ = format_dimension_row(k, v, "", self._node_type)
            tree.insert(
                sec_p,
                tk.END,
                values=(sym, meaning, "", format_param_value(k, v)),
                tags=("param",),
            )

        # ── 尺寸 ──
        if sol.dimensions:
            sec_d = tree.insert(
                "",
                tk.END,
                values=("", "── 构筑物尺寸 ──", "", ""),
                tags=("section",),
                open=True,
            )
            for k, (v, u) in sol.dimensions.items():
                sym, meaning, default_u = format_dimension_row(k, v, u, self._node_type)
                unit = u or default_u
                tree.insert(
                    sec_d,
                    tk.END,
                    values=(sym, meaning, unit, self._fmt_val(v)),
                    tags=("dim",),
                )

        # ── 约束校核 ──
        check_lines = []
        if sol.checks:
            for cn, (passed, actual, limit, unit) in sol.checks.items():
                mark = "✓" if passed else "✗"
                check_lines.append(f"[{mark}] {cn}: {actual}")
        if sol.cost_wan_yuan > 0:
            check_lines.append(f"概算成本: {sol.cost_wan_yuan:.1f} 万元")
        if check_lines:
            self._detail_check_text.insert("1.0", " | ".join(check_lines))

        # 样式
        tree.tag_configure(
            "section",
            background="#252525",
            foreground="#ffaa44",
            font=("Microsoft YaHei", 9, "bold"),
        )
        tree.tag_configure("param", foreground="#aaccff")
        tree.tag_configure("dim", foreground="#ddd")

    def _clear_detail(self):
        """清空详情面板"""
        self._detail_tree.delete(*self._detail_tree.get_children())
        self._detail_check_text.delete("1.0", tk.END)

    @staticmethod
    def _fmt_val(v) -> str:
        if isinstance(v, float):
            return f"{v:.2f}"
        return str(v)

    # ═══════════════════ 操作 ═══════════════════

    def _select_applied(self):
        """自动选中与当前节点参数匹配的方案;无匹配则自动应用推荐方案并选中"""
        if getattr(self, "_selecting", False):
            return
        self._selecting = True
        try:
            applied_params = self._get_applied_params()
            children = self._tree.get_children()
            if not children:
                return

            # 优先找已标记为 applied 的行
            for iid in children:
                tags = self._tree.item(iid, "tags")
                if tags and "applied" in tags:
                    self._tree.selection_set(iid)
                    self._tree.see(iid)
                    self._tree.focus(iid)
                    self._on_row_selected(None)
                    return

            # 无匹配 → 自动应用首行方案并通知外部回调
            filtered = self._get_filtered_solutions()
            if filtered:
                idx = int(children[0])
                if 0 <= idx < len(filtered):
                    self._selected_idx = self._solutions.index(filtered[idx])
                    self._apply_current()
                    self._refresh_table()
                    if self._apply_callback:
                        self._apply_callback(self._solutions, self._selected_idx)
                    for iid in self._tree.get_children():
                        tags = self._tree.item(iid, "tags")
                        if tags and "applied" in tags:
                            self._tree.selection_set(iid)
                            self._tree.see(iid)
                            self._tree.focus(iid)
                            self._on_row_selected(None)
                            return
            self._select_recommended()
        finally:
            self._selecting = False

    def _select_recommended(self):
        """选中推荐方案(当前排序模式下排第一的可见方案)"""
        children = self._tree.get_children()
        if not children:
            return
        first_iid = children[0]
        self._tree.selection_set(first_iid)
        self._tree.see(first_iid)
        self._tree.focus(first_iid)
        # 从可见列表中找对应方案
        try:
            idx = int(first_iid)
            filtered = self._get_filtered_solutions()
            if 0 <= idx < len(filtered):
                sol = filtered[idx]
                self._selected_idx = self._solutions.index(sol)
                self._show_detail(sol)
        except (ValueError, IndexError):
            pass

    # ═══════════════ 查询/获取 ═══════════════
    def _get_filtered_solutions(self) -> list:
        """获取当前筛选+排序后的方案列表"""
        filtered = self._solutions
        for key, val in self._filters.items():
            if val is not None:
                filtered = [
                    s for s in filtered if abs(s.params.get(key, 0) - val) < 1e-9
                ]
        if self._sort_mode == "cost":
            return sorted(
                filtered,
                key=lambda s: s.cost_wan_yuan if s.cost_wan_yuan > 0 else float("inf"),
            )
        else:
            return sorted(filtered, key=lambda s: -s.robustness)

    def _apply_current(self):
        """将当前选中的方案写入后端节点(不刷新表格,供内部调用)

        关键修复: 应用方案后运行标量 execute() 验证, 避免向量化/标量双路径
        结果不一致导致导出时出现"硬约束未通过"的虚假告警.
        """
        if not self._backend_node or self._selected_idx < 0:
            return False
        sol = self._solutions[self._selected_idx]
        for key, value in sol.params.items():
            try:
                self._backend_node.set_param(key, value)
            except Exception as e:
                _log.warning("operation failed: %s", e, exc_info=True)

        from models.base import NodeResult, NodeState, WaterQuality, WaterFlow

        # ── 标量验证: 用真实的 calculate() 跑一遍, 作为 ground truth ──
        scalar_passed = True
        scalar_warnings = []
        if self._flow and self._quality:
            try:
                scalar_result, _, _ = self._backend_node.execute(
                    self._flow, self._quality
                )
                if scalar_result and scalar_result.success:
                    # 对比向量化 checks: 标量 Fail 但向量化 Pass 的 → 警告
                    for cn, (v_passed, v_actual, v_limit, v_unit) in sol.checks.items():
                        s_check = scalar_result.checks.get(cn)
                        if s_check is not None:
                            s_passed = s_check[0]
                            if v_passed and not s_passed:
                                scalar_passed = False
                                scalar_warnings.append(cn)
                    if not scalar_passed:
                        _log.warning(
                            "[方案验证] %s 标量与向量化结果不一致, 约束失败: %s",
                            self._node_type,
                            ", ".join(scalar_warnings),
                        )
                    # 使用标量结果 (ground truth) 覆盖向量化近似
                    self._backend_node._result = scalar_result
                    self._backend_node.state = NodeState.CLEAN
                    return True
                else:
                    scalar_passed = False
                    _log.warning(
                        "[方案验证] %s 标量计算失败, 回退到向量化结果",
                        self._node_type,
                    )
            except Exception as e:
                scalar_passed = False
                _log.warning(
                    "[方案验证] %s 标量验证异常: %s",
                    self._node_type,
                    e,
                )

        # ── 回退: 标量验证失败时使用向量化结果, 标记 DIRTY 强制 F5 重算 ──
        result = NodeResult(
            success=True, params=dict(sol.params), robustness=sol.robustness
        )
        for k, (v, u) in sol.dimensions.items():
            result.add_dimension(k, v, u)
        for cn, (passed, actual, limit, unit) in sol.checks.items():
            result.add_check(cn, passed, actual, limit, unit)
        if self._quality and any(
            getattr(self._quality, a, 0) > 0 for a in ["BOD5", "COD"]
        ):
            result.inlet_quality = WaterQuality(
                BOD5=self._quality.BOD5,
                COD=self._quality.COD,
                SS=self._quality.SS,
                NH3N=self._quality.NH3N,
                TN=self._quality.TN,
                TP=self._quality.TP,
                pH=getattr(self._quality, "pH", 7.0),
            )
            rates = self._backend_node.get_removal_rates()
            result.outlet_quality = self._quality.apply_removal(rates)
        self._backend_node._result = result
        # 标量验证未通过 → 标记 DIRTY 确保 F5 重算, 同时保留向量化结果供参考
        if not scalar_passed:
            self._backend_node.state = NodeState.DIRTY
            result.add_warning(
                f"向量化/标量结果不一致, 约束差异: {', '.join(scalar_warnings)}"
                if scalar_warnings
                else "标量验证未通过, 请按 F5 重新计算"
            )
        else:
            self._backend_node.state = NodeState.CLEAN
        return True

    # ═══════════════ 事件回调 ═══════════════
    def _on_apply(self):
        """应用选中方案到后端节点 — 同时将向量化结果存入缓存, 避免导出时标量重算偏差"""
        if self._apply_current():
            self._refresh_table()
            self._select_applied()
            if self._apply_callback:
                self._apply_callback(self._solutions, self._selected_idx)

    # ═══════════════ 面板刷新 ═══════════════
    def _on_refresh(self):
        """重新枚举方案空间(强制忽略缓存)"""
        if self._backend_node and self._flow and self._quality:
            self.load_node(
                self._backend_node, self._flow, self._quality, force_recompute=True
            )

    # ═══════════════ 导出/输出 ═══════════════
    def _export_to_excel(self) -> None:
        """导出所有可行方案到 Excel 文件"""
        from tkinter import filedialog, messagebox

        if not self._solutions:
            messagebox.showinfo("提示", "暂无可行方案可导出")
            return

        path = filedialog.asksaveasfilename(
            title="导出方案到Excel",
            defaultextension=".xlsx",
            filetypes=[("Excel文件", "*.xlsx")],
            initialfile=f"{self._node_type}_方案列表.xlsx",
        )
        if not path:
            return

        try:
            import openpyxl
            from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
        except ImportError:
            messagebox.showerror(
                "错误", "需要 openpyxl 库,请运行: pip install openpyxl"
            )
            return

        wb = openpyxl.Workbook()

        # ── Sheet 1: 方案总览 ──
        ws = wb.active
        ws.title = "方案总览"

        header_font = Font(bold=True, size=11)
        header_fill = PatternFill(
            start_color="4472C4", end_color="4472C4", fill_type="solid"
        )
        header_font_white = Font(bold=True, size=11, color="FFFFFF")
        thin_border = Border(
            left=Side(style="thin"),
            right=Side(style="thin"),
            top=Side(style="thin"),
            bottom=Side(style="thin"),
        )

        # 收集所有参数键
        param_keys = list(self._solutions[0].params.keys()) if self._solutions else []
        # 安全排序列
        rank_keys = ["robustness", "n_constraints_pass", "cost"]

        headers = ["排名"] + rank_keys + param_keys
        for col_idx, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col_idx, value=header)
            cell.font = header_font_white
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal="center")
            cell.border = thin_border

        for row_idx, sol in enumerate(self._solutions, 2):
            ws.cell(row=row_idx, column=1, value=row_idx - 1).border = thin_border
            ws.cell(row=row_idx, column=2, value=round(sol.robustness, 3)).border = (
                thin_border
            )
            ws.cell(row=row_idx, column=3, value=sol.n_constraints_pass).border = (
                thin_border
            )
            ws.cell(row=row_idx, column=4, value=round(sol.cost, 1)).border = (
                thin_border
            )
            for j, key in enumerate(param_keys):
                val = sol.params.get(key, "")
                cell = ws.cell(row=row_idx, column=5 + j, value=val)
                cell.border = thin_border

        ws.column_dimensions["A"].width = 8
        for col_idx in range(2, len(headers) + 1):
            ws.column_dimensions[openpyxl.utils.get_column_letter(col_idx)].width = 14

        # ── Sheet 2: 详细参数 ──
        ws2 = wb.create_sheet("详细参数")
        dim_keys = []
        if self._solutions and hasattr(self._solutions[0], "dimensions"):
            dim_keys = list(self._solutions[0].dimensions.keys())

        headers2 = ["排名"] + param_keys + dim_keys
        for col_idx, header in enumerate(headers2, 1):
            cell = ws2.cell(row=1, column=col_idx, value=header)
            cell.font = header_font_white
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal="center")
            cell.border = thin_border

        for row_idx, sol in enumerate(self._solutions, 2):
            ws2.cell(row=row_idx, column=1, value=row_idx - 1).border = thin_border
            for j, key in enumerate(param_keys):
                ws2.cell(
                    row=row_idx, column=2 + j, value=sol.params.get(key, "")
                ).border = thin_border
            if hasattr(sol, "dimensions"):
                for j, key in enumerate(dim_keys):
                    ws2.cell(
                        row=row_idx,
                        column=2 + len(param_keys) + j,
                        value=sol.dimensions.get(key, ""),
                    ).border = thin_border

        # ── Sheet 3: 校核结果 ──
        if self._solutions and hasattr(self._solutions[0], "checks"):
            ws3 = wb.create_sheet("校核结果")
            check_keys = list(self._solutions[0].checks.keys())
            headers3 = ["排名"] + check_keys
            for col_idx, header in enumerate(headers3, 1):
                cell = ws3.cell(row=1, column=col_idx, value=header)
                cell.font = header_font_white
                cell.fill = header_fill
                cell.alignment = Alignment(horizontal="center")
                cell.border = thin_border

            for row_idx, sol in enumerate(self._solutions, 2):
                ws3.cell(row=row_idx, column=1, value=row_idx - 1).border = thin_border
                for j, key in enumerate(check_keys):
                    passed = sol.checks.get(key, False)
                    cell = ws3.cell(
                        row=row_idx, column=2 + j, value="✓" if passed else "✗"
                    )
                    cell.border = thin_border
                    cell.fill = PatternFill(
                        start_color="C6EFCE" if passed else "FFC7CE",
                        end_color="C6EFCE" if passed else "FFC7CE",
                        fill_type="solid",
                    )
                    cell.alignment = Alignment(horizontal="center")

        try:
            wb.save(path)
            messagebox.showinfo(
                "导出完成", f"已导出 {len(self._solutions)} 个方案到:\n{path}"
            )
        except Exception as e:
            messagebox.showerror("导出失败", str(e))

    # ═══════════════ 查询/获取 ═══════════════
    def _get_applied_params(self) -> Dict[str, float]:
        """获取后端节点当前应用的参数(仅当节点为 CLEAN 时有效)"""
        if not self._backend_node or self._backend_node.is_dirty:
            return {}
        return {
            pd.key: pd.value
            for pd in self._backend_node.get_param_defs()
            if pd.value > 0 or pd.key in self._backend_node._params
        }

    # ═══════════════ 事件回调 ═══════════════
    def _is_solution_applied(self, sol: Solution, applied: Dict[str, float]) -> bool:
        """检查方案是否已被应用到后端节点"""
        if not applied:
            return False
        # 比较自由变量参数 (来自 discretization config)
        for key in sol.params:
            if key in applied:
                if abs(sol.params[key] - applied[key]) > 1e-6:
                    return False
        return True

    def _clear_all(self):
        """清除所有内容"""
        for item in self._tree.get_children():
            self._tree.delete(item)
        self._clear_detail()
        for w in self._filter_frame.winfo_children():
            w.destroy()
        self._selected_idx = -1
        self._solutions = []
        self._filters = {}
